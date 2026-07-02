import joblib
import random
import os
import json
import warnings
import mysql.connector
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -- Load external config (config.json, same folder as this script) ----------
CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")

DEFAULT_CONFIG = {
    "database": {
        "host": "127.0.0.1",
        "port": 3306,
        "user": "acore",
        "password": "acore",
        "database": "acore_world"
    },
    "entry_id_start": 91000,
    "brain_file": "blizzlike_master_brain.pkl",
    "material_library_file": "material_library.joblib"
}

def load_config(path):
    if not os.path.exists(path):
        print(f"⚠️  No config.json found at {path}.")
        print("    Creating one with default values -- edit it with your DB credentials and rerun.")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(DEFAULT_CONFIG, f, indent=2)
        print(f"    Wrote default config to {path}")
        exit()
    try:
        with open(path, "r", encoding="utf-8") as f:
            cfg = json.load(f)
    except Exception as e:
        print(f"❌ Failed to parse {path}: {e}")
        exit()
    # Fill in any missing keys with defaults so partial configs still work
    for key, val in DEFAULT_CONFIG.items():
        if key not in cfg:
            cfg[key] = val
    for key, val in DEFAULT_CONFIG["database"].items():
        cfg["database"].setdefault(key, val)
    return cfg

CONFIG = load_config(CONFIG_PATH)

DB_CONFIG = {
    "user": CONFIG["database"]["user"],
    "password": CONFIG["database"]["password"],
    "host": CONFIG["database"]["host"],
    "database": CONFIG["database"]["database"],
    "port": CONFIG["database"]["port"]
}
ENTRY_ID_START = CONFIG["entry_id_start"]

material_library = joblib.load(CONFIG["material_library_file"])
warnings.filterwarnings("ignore")

try:
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()
    print(f"✅ Connected to {DB_CONFIG['database']} database ({DB_CONFIG['host']}:{DB_CONFIG['port']}).")
except Exception as e:
    print(f"❌ Database connection failed: {e}")
    exit()
try:
    master = joblib.load(CONFIG["brain_file"])
    lookup_database = master["lookup_database"]
    slot_budget_curves  = master.get("slot_budget_curves", {})
    global_budget_curves = master["global_budget_curves"]
    archetype_profiles = master["archetype_profiles"]
    name_database = master["name_database"]
    subclass_delays = master.get("subclass_delays", {})
    weapon_nouns = master.get("weapon_nouns", {})
    armor_nouns = master.get("armor_nouns", {})
except Exception as e:
    print("❌ Critical Error: Model components failed to load. Run train_brain.py first.")
    exit()

# -- Load valid RandomProperty / RandomSuffix enchantment template entries -----
# item_enchantment_template.entry is what goes into item_template.RandomProperty
# or item_template.RandomSuffix.  We read all DISTINCT entries from the live DB
# so the generator always picks from entries that actually exist.
#
# RandomProperty: items below level 20 -- prefix names ("of the Bear" etc.),
#                 flat bonuses from ItemRandomProperties.dbc
# RandomSuffix:   items level 20+ -- suffix names ("of Stamina" etc.),
#                 scaled by allocation points via RandomPropertiesPoints.dbc
#
# Both columns cannot be non-zero on the same item (AC hard constraint).
try:
    _rand_conn   = get_db_connection()
    _rand_cursor = _rand_conn.cursor()
    _rand_cursor.execute("SELECT DISTINCT entry FROM item_enchantment_template ORDER BY entry")
    _ALL_ENCHANT_ENTRIES = [row[0] for row in _rand_cursor.fetchall()]
    _rand_cursor.close()
    _rand_conn.close()
    print(f"  Loaded {len(_ALL_ENCHANT_ENTRIES)} item_enchantment_template entries for RandomProp/Suf.")
except Exception as _e:
    print(f"  WARNING: Could not load item_enchantment_template ({_e}). RandomProp/Suf disabled.")
    _ALL_ENCHANT_ENTRIES = []

# Global budget nerf applied to every generated item after fuzz.
# 1.0 = no nerf, 0.90 = 10% nerf, 0.85 = 15% nerf, etc.
GLOBAL_BUDGET_NERF = 0.90

STAT_NAMES = {
    0: ("Mana", "Primary/Resource"), 1: ("Health", "Primary/Resource"), 3: ("Agility", "Primary"),
    4: ("Strength", "Primary"), 5: ("Intellect", "Primary"), 6: ("Spirit", "Primary"), 7: ("Stamina", "Primary"),
    12: ("Defense Rating", "Secondary"), 13: ("Dodge Rating", "Secondary"), 14: ("Parry Rating", "Secondary"),
    15: ("Block Rating", "Secondary"), 16: ("Melee Hit Rating", "Secondary"), 17: ("Ranged Hit Rating", "Secondary"),
    18: ("Spell Hit Rating", "Secondary"), 19: ("Melee Crit Rating", "Secondary"), 20: ("Ranged Crit Rating", "Secondary"),
    21: ("Spell Crit Rating", "Secondary"), 28: ("Melee Haste Rating", "Secondary"), 30: ("Spell Haste Rating", "Secondary"),
    31: ("Hit Rating", "Secondary"), 32: ("Critical Strike Rating", "Secondary"), 35: ("Resilience", "Secondary"),
    36: ("Haste Rating", "Secondary"), 37: ("Expertise Rating", "Secondary"), 38: ("Attack Power", "Secondary"),
    43: ("Mana Regeneration (MP5)", "Secondary"), 44: ("Armor Penetration Rating", "Secondary"),
    45: ("Spell Power", "Secondary"), 48: ("Block Value", "Secondary")
}

# Updated master catalog with 1H Mace added (ID 22)
CATEGORIES = {
    # Weapons
    1: {"name": "Wand", "class": 2, "subclass": 19, "InventoryType": 26, "delay": 1800, "dmg_type1": 3},
    2: {"name": "One-Handed Sword", "class": 2, "subclass": 7, "InventoryType": 13, "delay": 2600, "dmg_type1": 0},
    3: {"name": "Two-Handed Sword", "class": 2, "subclass": 8, "InventoryType": 17, "delay": 3600, "dmg_type1": 0},
    4: {"name": "Dagger", "class": 2, "subclass": 15, "InventoryType": 13, "delay": 1800, "dmg_type1": 0},
    5: {"name": "Staff", "class": 2, "subclass": 10, "InventoryType": 17, "delay": 3200, "dmg_type1": 0},
    6: {"name": "Fist Weapon", "class": 2, "subclass": 13, "InventoryType": 13, "delay": 2600, "dmg_type1": 0},
    7: {"name": "One-Handed Axe", "class": 2, "subclass": 0, "InventoryType": 13, "delay": 2600, "dmg_type1": 0},
    8: {"name": "Two-Handed Mace", "class": 2, "subclass": 5, "InventoryType": 17, "delay": 3600, "dmg_type1": 0},
    9: {"name": "Polearm", "class": 2, "subclass": 6, "InventoryType": 17, "delay": 3500, "dmg_type1": 0},
    10: {"name": "Two-Handed Axe", "class": 2, "subclass": 1, "InventoryType": 17, "delay": 3600, "dmg_type1": 0},
    11: {"name": "Crossbow", "class": 2, "subclass": 18, "InventoryType": 26, "delay": 3000, "dmg_type1": 0},
    12: {"name": "Bow", "class": 2, "subclass": 2, "InventoryType": 15, "delay": 2800, "dmg_type1": 0},
    13: {"name": "Gun", "class": 2, "subclass": 3, "InventoryType": 26, "delay": 2800, "dmg_type1": 0},
    14: {"name": "Thrown Weapon", "class": 2, "subclass": 16, "InventoryType": 25, "delay": 2000, "dmg_type1": 0},
    15: {"name": "One-Handed Mace", "class": 2, "subclass": 4, "InventoryType": 13, "delay": 2600, "dmg_type1": 0},
    
    # Cloth (Subclass 1)
    16: {"name": "Cloth Helm", "class": 4, "subclass": 1, "InventoryType": 1, "delay": 0, "dmg_type1": 0},
    17: {"name": "Cloth Shoulders", "class": 4, "subclass": 1, "InventoryType": 3, "delay": 0, "dmg_type1": 0},
    18: {"name": "Cloth Chest", "class": 4, "subclass": 1, "InventoryType": 5, "delay": 0, "dmg_type1": 0},
    19: {"name": "Cloth Wrist", "class": 4, "subclass": 1, "InventoryType": 9, "delay": 0, "dmg_type1": 0},
    20: {"name": "Cloth Gloves", "class": 4, "subclass": 1, "InventoryType": 10, "delay": 0, "dmg_type1": 0},
    21: {"name": "Cloth Waist", "class": 4, "subclass": 1, "InventoryType": 6, "delay": 0, "dmg_type1": 0},
    22: {"name": "Cloth Legs", "class": 4, "subclass": 1, "InventoryType": 7, "delay": 0, "dmg_type1": 0},
    23: {"name": "Cloth Feet", "class": 4, "subclass": 1, "InventoryType": 8, "delay": 0, "dmg_type1": 0},
    
    # Leather (Subclass 2)
    24: {"name": "Leather Helm", "class": 4, "subclass": 2, "InventoryType": 1, "delay": 0, "dmg_type1": 0},
    25: {"name": "Leather Shoulders", "class": 4, "subclass": 2, "InventoryType": 3, "delay": 0, "dmg_type1": 0},
    26: {"name": "Leather Chest", "class": 4, "subclass": 2, "InventoryType": 5, "delay": 0, "dmg_type1": 0},
    27: {"name": "Leather Wrist", "class": 4, "subclass": 2, "InventoryType": 9, "delay": 0, "dmg_type1": 0},
    28: {"name": "Leather Gloves", "class": 4, "subclass": 2, "InventoryType": 10, "delay": 0, "dmg_type1": 0},
    29: {"name": "Leather Waist", "class": 4, "subclass": 2, "InventoryType": 6, "delay": 0, "dmg_type1": 0},
    30: {"name": "Leather Legs", "class": 4, "subclass": 2, "InventoryType": 7, "delay": 0, "dmg_type1": 0},
    31: {"name": "Leather Feet", "class": 4, "subclass": 2, "InventoryType": 8, "delay": 0, "dmg_type1": 0},
    
    # Mail (Subclass 3)
    32: {"name": "Mail Helm", "class": 4, "subclass": 3, "InventoryType": 1, "delay": 0, "dmg_type1": 0},
    33: {"name": "Mail Shoulders", "class": 4, "subclass": 3, "InventoryType": 3, "delay": 0, "dmg_type1": 0},
    34: {"name": "Mail Chest", "class": 4, "subclass": 3, "InventoryType": 5, "delay": 0, "dmg_type1": 0},
    35: {"name": "Mail Wrist", "class": 4, "subclass": 3, "InventoryType": 9, "delay": 0, "dmg_type1": 0},
    36: {"name": "Mail Gloves", "class": 4, "subclass": 3, "InventoryType": 10, "delay": 0, "dmg_type1": 0},
    37: {"name": "Mail Waist", "class": 4, "subclass": 3, "InventoryType": 6, "delay": 0, "dmg_type1": 0},
    38: {"name": "Mail Legs", "class": 4, "subclass": 3, "InventoryType": 7, "delay": 0, "dmg_type1": 0},
    39: {"name": "Mail Feet", "class": 4, "subclass": 3, "InventoryType": 8, "delay": 0, "dmg_type1": 0},
    
    # Plate (Subclass 4)
    40: {"name": "Plate Helm", "class": 4, "subclass": 4, "InventoryType": 1, "delay": 0, "dmg_type1": 0},
    41: {"name": "Plate Shoulders", "class": 4, "subclass": 4, "InventoryType": 3, "delay": 0, "dmg_type1": 0},
    42: {"name": "Plate Chest", "class": 4, "subclass": 4, "InventoryType": 5, "delay": 0, "dmg_type1": 0},
    43: {"name": "Plate Wrist", "class": 4, "subclass": 4, "InventoryType": 9, "delay": 0, "dmg_type1": 0},
    44: {"name": "Plate Gloves", "class": 4, "subclass": 4, "InventoryType": 10, "delay": 0, "dmg_type1": 0},
    45: {"name": "Plate Waist", "class": 4, "subclass": 4, "InventoryType": 6, "delay": 0, "dmg_type1": 0},
    46: {"name": "Plate Legs", "class": 4, "subclass": 4, "InventoryType": 7, "delay": 0, "dmg_type1": 0},
    47: {"name": "Plate Feet", "class": 4, "subclass": 4, "InventoryType": 8, "delay": 0, "dmg_type1": 0},
    
    # Miscellaneous (Subclass 0)
    48: {"name": "Cloak", "class": 4, "subclass": 1, "InventoryType": 16, "delay": 0, "dmg_type1": 0},
    49: {"name": "Necklace", "class": 4, "subclass": 0, "InventoryType": 2, "delay": 0, "dmg_type1": 0},
    50: {"name": "Ring / Band", "class": 4, "subclass": 0, "InventoryType": 11, "delay": 0, "dmg_type1": 0},
    51: {"name": "Shield", "class": 4, "subclass": 6, "InventoryType": 14, "delay": 0, "dmg_type1": 0},
    52: {"name": "Offhand", "class": 4, "subclass": 0, "InventoryType": 23, "delay": 0, "dmg_type1": 0},
}

QUALITIES = {
    1: {"name": "Uncommon (Green)",          "code": 2,        "multi": False},
    2: {"name": "Rare (Blue)",               "code": 3,        "multi": False},
    3: {"name": "Epic (Purple)",             "code": 4,        "multi": False},
    4: {"name": "Legendary (Orange)",        "code": 5,        "multi": False},
    5: {"name": "Uncommon and Rare",         "code": [2, 3],   "multi": True},
    6: {"name": "Rare and Epic",             "code": [3, 4],   "multi": True},
    7: {"name": "Uncommon, Rare and Epic",   "code": [2, 3, 4],"multi": True},
}

BLUEPRINTS = {
    "AGI_DPS":   {"pool": [3, 7, 31, 32, 36, 38, 44], "anchors": [3], "weights": {3: 130, 7: 100, 31: 90, 32: 90, 36: 60, 38: 80, 44: 70}},
    "STR_DPS":   {"pool": [4, 7, 31, 32, 36, 38, 44], "anchors": [4], "weights": {4: 130, 7: 100, 31: 90, 32: 90, 36: 60, 38: 80, 44: 70}},
    "SP_DPS":    {"pool": [5, 6, 7, 45, 31, 32, 36],     "anchors": [5], "weights": {5: 120, 45: 130, 7: 90, 31: 85, 32: 85, 36: 80, 6: 70}},
    "HEALER":    {"pool": [5, 6, 7, 45, 32, 36, 43],   "anchors": [5], "weights": {5: 110, 6: 110, 45: 120, 7: 90, 32: 80, 36: 80, 43: 75}},
    "STR_TANK":  {"pool": [4, 7, 12, 13, 14, 31, 37],  "anchors": [7, 4],  "weights": {7: 130, 4: 100, 12: 85, 13: 90, 14: 90, 31: 75, 37: 85}},
    "AGI_TANK":  {"pool": [3, 7, 12, 13, 31, 37, 32, 36, 38], "anchors": [3, 7], "weights": {3: 150, 7: 140, 12: 85, 13: 100, 31: 80, 37: 85, 32: 45, 36: 45, 38: 40}},
    "AGI_INT_DPS": {"pool": [3, 5, 7, 31, 32, 36, 38, 44], "anchors": [3, 5], "weights": {3: 130, 5: 100, 7:100, 31: 90, 32: 90, 36: 60, 38: 80, 44: 70}}
}

# (Helper functions interpolate_macro_budget, interpolate_local_dps, get_interpolated_properties remain same)
def export_to_excel_tooltips(internal_memory, filename="generated_items_tooltips.xlsx"):
    """
    Generates an authentic dark-themed Excel sheet styling items as Blizzard tooltips
    by dynamically tracking and outlining each item card.
    """
    if not internal_memory:
        print("⚠️ Core collection cache array is empty. No Excel file created.")
        return

    # Excel hard-caps a worksheet at 1,048,576 rows. Mass Creation can easily
    # produce enough items to blow past that (each item block uses up to ~18
    # rows), which used to crash openpyxl mid-write and abort the export
    # before wb.save() ever ran -- so the SQL/CSV files would appear but the
    # tooltips workbook never got saved. Instead of capping the data, we spill
    # over into additional worksheets within the SAME workbook/file once a
    # sheet gets close to the limit, so every item is still represented no
    # matter how large the run is.
    EXCEL_MAX_ROW = 1_048_576
    MAX_ROWS_PER_ITEM = 20          # worst-case rows a single item block can consume
    SHEET_ROW_CUTOFF = EXCEL_MAX_ROW - MAX_ROWS_PER_ITEM - 5  # safety margin

    def setup_sheet(target_ws):
        target_ws.views.sheetView[0].showGridLines = True
        target_ws.column_dimensions['A'].width = 3
        target_ws.column_dimensions['B'].width = 16
        target_ws.column_dimensions['C'].width = 14
        target_ws.column_dimensions['D'].width = 14
        target_ws.column_dimensions['E'].width = 16

    wb = Workbook()
    ws = wb.active
    sheet_num = 1
    ws.title = f"Item Tooltips {sheet_num}"
    setup_sheet(ws)

    # Hex mapping for Blizzard item quality font strings
    QUALITY_COLORS = {
        0: "9D9D9D",  # Poor / Grey
        1: "FFFFFF",  # Common / White
        2: "1EFF00",  # Uncommon / Green
        3: "0070DD",  # Rare / Blue
        4: "A335EE",  # Epic / Purple
        5: "FF8000"   # Legendary / Orange
    }
    
    # Human-readable InventoryType Slot tags
    SLOT_NAMES = {
        1: "Head", 2: "Neck", 3: "Shoulder", 5: "Chest", 6: "Waist", 7: "Legs", 
        8: "Feet", 9: "Wrist", 10: "Hands", 11: "Finger", 12: "Trinket", 
        13: "One-Hand", 14: "Shield", 15: "Ranged", 16: "Back", 17: "Two-Hand", 
        20: "Chest", 21: "Main-Hand", 22: "Off-Hand", 23: "Held In Off-Hand", 
        25: "Thrown", 26: "Ranged"
    }
    
    # Subclass text parser mappings
    SUBCLASS_NAMES = {
        2: {0: "Axe", 1: "Two-Handed Axe", 2: "Bow", 3: "Gun", 4: "Mace", 5: "Two-Handed Mace", 6: "Polearm", 7: "Sword", 8: "Two-Handed Sword", 10: "Staff", 13: "Fist Weapon", 15: "Dagger", 16: "Thrown", 18: "Crossbow", 19: "Wand"},
        4: {0: "Miscellaneous", 1: "Cloth", 2: "Leather", 3: "Mail", 4: "Plate", 6: "Shield"}
    }

    bg_fill = PatternFill(start_color="111216", end_color="111216", fill_type="solid")
    thin_grey = Side(border_style="thin", color="3A3F4D")
    
    current_row = 2  # Start row coordinate position
    total_items = len(internal_memory)

    for item_idx, item in enumerate(internal_memory):
        # Spill into a fresh worksheet if this item's block could push us
        # past Excel's row ceiling on the current sheet.
        if current_row > SHEET_ROW_CUTOFF:
            sheet_num += 1
            ws = wb.create_sheet(title=f"Item Tooltips {sheet_num}")
            setup_sheet(ws)
            current_row = 2

        if total_items > 2000 and (item_idx % 2000 == 0 or item_idx == total_items - 1):
            print(f"\r  📦 Writing tooltip {item_idx + 1:,}/{total_items:,} "
                  f"(sheet {sheet_num})...", end="", flush=True)

        c = item["config"]
        q_code = item.get("quality", 2)
        q_color = QUALITY_COLORS.get(q_code, "FFFFFF")
        
        start_row = current_row  # Track where this specific item box begins
        
        # Helper inner function to write background-painted rows smoothly.
        # NOTE: we intentionally do NOT call ws.merge_cells() here. openpyxl's
        # merge-range bookkeeping gets quadratically slower as the number of
        # merged ranges grows, which is what made large Mass Creation exports
        # take an effectively unbounded amount of time (and look "broken").
        # Left-aligned text overflows visually into empty neighboring cells
        # in Excel anyway, so the look is preserved without the cost.
        def write_row(val_left, val_right=None, font_left=None, font_right=None, merge_all=False):
            nonlocal current_row
            for col in range(2, 6):
                ws.cell(row=current_row, column=col).fill = bg_fill
                
            if merge_all:
                cell = ws.cell(row=current_row, column=2, value=val_left)
                if font_left: cell.font = font_left
            else:
                cell_l = ws.cell(row=current_row, column=2, value=val_left)
                if font_left: cell_l.font = font_left
                
                # Placed in the rightmost column so right-aligned overflow
                # still lands flush against the same edge the old D:E merge did.
                cell_r = ws.cell(row=current_row, column=5, value=val_right)
                if font_right: cell_r.font = font_right
                cell_r.alignment = Alignment(horizontal="right")
            current_row += 1

        # 1. Item Name Header
        write_row(item["name"], font_left=Font(name="Calibri", size=13, bold=True, color=q_color), merge_all=True)
        
        # 2. Item Level Line
        write_row(f"Item Level {item['ilvl']}", font_left=Font(name="Calibri", size=10, bold=True, color="FFD100"), merge_all=True)
        
        # 3. Binding Text Rule
        _bonding  = item.get("bonding", 1 if q_code >= 3 else 2)
        bind_text = "Binds when picked up" if _bonding == 1 else "Binds when equipped"
        write_row(bind_text, font_left=Font(name="Calibri", size=10, color="FFFFFF"), merge_all=True)
        
        # 4. Slot Type vs Equipment Class Split Line
        slot_str = SLOT_NAMES.get(c["InventoryType"], "Equippable")
        sub_str = SUBCLASS_NAMES.get(c["class"], {}).get(c["subclass"], "")
        write_row(slot_str, sub_str, font_left=Font(name="Calibri", size=10, color="FFFFFF"), font_right=Font(name="Calibri", size=10, color="FFFFFF"))
        
        # 5. Weapon Combat Metrics Panel Split
        if c["class"] == 2:
            dmg_str = f"{item['dmg_min']} - {item['dmg_max']} Damage"
            speed_str = f"Speed {item['delay']/1000:.2f}"
            write_row(dmg_str, speed_str, font_left=Font(name="Calibri", size=10, color="FFFFFF"), font_right=Font(name="Calibri", size=10, color="FFFFFF"))
            
            dps_str = f"({item['dps']:.1f} damage per second)"
            write_row(dps_str, font_left=Font(name="Calibri", size=10, color="FFFFFF"), merge_all=True)
            
        # 6. Armor Mitigation Value Panel
        elif c["class"] == 4 and item.get("armor", 0) > 0:
            write_row(f"{item['armor']} Armor", font_left=Font(name="Calibri", size=10, color="FFFFFF"), merge_all=True)
            
        # 7. Attributes Allocation Parser (Forcing Primary over Secondary)
        stats_data = item.get("stats", {})
        processed_stats = []
        
        if isinstance(stats_data, dict):
            if "stat_type1" in stats_data:
                for i in range(1, 11):
                    st = stats_data.get(f"stat_type{i}", 0)
                    sv = stats_data.get(f"stat_value{i}", 0)
                    if st != 0 and sv != 0: processed_stats.append((st, sv))
            else:
                for st, sv in stats_data.items():
                    if int(st) != 0 and int(sv) != 0: processed_stats.append((int(st), int(sv)))

        # Separate into prioritized display buckets based on text color rules
        white_stats = []
        green_stats = []

        for s_type, s_val in processed_stats:
            if s_type in STAT_NAMES:
                stat_name, stat_group = STAT_NAMES[s_type]
                stat_color = "00FF00" if stat_group == "Secondary" else "FFFFFF"
                
                # Format textual rules ahead of time
                if stat_group == "Secondary" and s_type != 38: # Attack Power behaves like a primary text layout
                    display_value = f"Equip: Increases {stat_name} by {s_val}."
                else:
                    display_value = f"+{s_val} {stat_name}"
                
                # Append to respective bucket
                if stat_color == "FFFFFF":
                    white_stats.append((display_value, stat_color))
                else:
                    green_stats.append((display_value, stat_color))

        # Render White Primaries completely before Green Secondaries
        for display_value, color in (white_stats + green_stats):
            write_row(display_value, font_left=Font(name="Calibri", size=10, color=color), merge_all=True)

        # 8. Durability Frame Line
        write_row("Durability 120 / 120", font_left=Font(name="Calibri", size=10, color="FFFFFF"), merge_all=True)
        
        # 9. Character Level Prerequisite Restrictions
        write_row(f"Requires Level {item['required_level']}", font_left=Font(name="Calibri", size=10, color="FFFFFF"), merge_all=True)
        
        # (Note: Step 10 'Chance on hit' lore block removed from sequence successfully)
            
        # 11. Sell Price Node Breakdown Matrix (Gold, Silver, Copper Calculations)
        raw_copper = item.get("sell_price", 0)
        gold = int(raw_copper // 10000)
        silver = int((raw_copper % 10000) // 100)
        copper = int(raw_copper % 100)
        
        price_str = "Sell Price: "
        if gold > 0: price_str += f"{gold}g "
        if silver > 0 or gold > 0: price_str += f"{silver}s "
        price_str += f"{copper}c"
        write_row(price_str, font_left=Font(name="Calibri", size=10, color="FFFFFF"), merge_all=True)
        
        # 12. Draw outer bounding borders around the item container block
        for r in range(start_row, current_row):
            for col in range(2, 6):
                cell = ws.cell(row=r, column=col)
                cell.border = Border(
                    left=thin_grey if col == 2 else None,
                    right=thin_grey if col == 5 else None,
                    top=thin_grey if r == start_row else None,
                    bottom=thin_grey if r == current_row - 1 else None
                )
                
        current_row += 3  # Insert spacing margin rows between rendered card boxes

    if total_items > 2000:
        print()  # close out the progress line

    wb.save(filename)
    sheet_note = f" across {sheet_num} sheets" if sheet_num > 1 else ""
    print(f"📦 Excel Tooltip Sheet successfully written to system workspace: {filename}{sheet_note}")
def get_dynamic_sell_price(sheet, ilvl):
    """
    Finds the closest item level nodes in the sheet and returns the 
    linearly interpolated average sell price in copper.
    """
    if not sheet:
        return 0
    if ilvl <= sheet[0]["itemlevel"]:
        return sheet[0]["avg_sell_price"]
    if ilvl >= sheet[-1]["itemlevel"]:
        return sheet[-1]["avg_sell_price"]
        
    for i in range(len(sheet) - 1):
        p1 = sheet[i]
        p2 = sheet[i+1]
        if p1["itemlevel"] <= ilvl <= p2["itemlevel"]:
            if p2["itemlevel"] == p1["itemlevel"]:
                return p1["avg_sell_price"]
            ratio = (ilvl - p1["itemlevel"]) / (p2["itemlevel"] - p1["itemlevel"])
            return p1["avg_sell_price"] + ratio * (p2["avg_sell_price"] - p1["avg_sell_price"])
    return sheet[0]["avg_sell_price"]

def calculate_item_sell_price(lookup_database, category, quality_code, ilvl):
    """
    Finds a sell price baseline using a rigid multi-tiered fallback architecture:
    1. Exact match (Quality, Slot, Subclass)
    2. Alternative Quality match (Same Slot & Subclass)
    3. Global Subclass match (Any Slot or Quality matching the base weapon/armor type)
    """
    cls, subcls, inv_type = category["class"], category["subclass"], category["InventoryType"]
    
    # Tier 1: Exact Match
    exact_key = (cls, subcls, inv_type, quality_code)
    if exact_key in lookup_database and lookup_database[exact_key]:
        return get_dynamic_sell_price(lookup_database[exact_key], ilvl)
        
    # Tier 2: Check alternative qualities for the same slot/subclass configuration
    quality_ladder = [quality_code] + [3, 4, 2, 1, 0, 5] # Priorities: Rare, Epic, Uncommon, etc.
    for q in quality_ladder:
        alt_key = (cls, subcls, inv_type, q)
        if alt_key in lookup_database and lookup_database[alt_key]:
            return get_dynamic_sell_price(lookup_database[alt_key], ilvl)
            
    # Tier 3: Broadest Subclass Fallback (Scan full database for matching class/subclass structure)
    for (c, sc, it, q), sheet in lookup_database.items():
        if c == cls and sc == subcls and sheet:
            return get_dynamic_sell_price(sheet, ilvl)
            
    return 500 # 5 Silver safety baseline if completely empty
def get_weapon_delay(category_data):
    """
    Fetches the database-calculated average speed for a weapon subclass,
    applies a +/-15% swing variation, and snaps it cleanly to the nearest 100ms.
    """
    subclass = category_data["subclass"]
    
    # Grab the true average from the DBC/Database via train_brain. 
    # Fallback to the hardcoded CATEGORIES delay if the dataset doesn't have it.
    base_delay = subclass_delays.get(subclass, category_data["delay"])
    
    # Calculate 15% boundaries
    min_delay = base_delay * 0.85
    max_delay = base_delay * 1.15
    
    # Roll the randomized raw speed value
    raw_delay = random.uniform(min_delay, max_delay)
    
    # Snap smoothly to the nearest 100ms step
    return int(round(raw_delay / 100) * 100)
def get_sheathe_type(cat):
    """
    Returns the Sheath ID based on WoW standards:
    1: 2H Weapon, 2: Staff, 3: 1H Weapon, 4: Shield, 7: Off-hand
    """
    inv = cat.get("InventoryType")
    cls = cat.get("class")
    sub = cat.get("subclass")

    # Shield (InvType 6)
    if inv == 6: return 4
    
    # Armor (Class 4) - Usually no sheath type
    if cls == 4: return 0
    
    # Weapons (Class 2)
    if cls == 2:
        # Staff (Subclass 10)
        if sub == 10: return 2
        # Off-hand (InvType 23)
        if inv == 23: return 7
        # 2H Weapons (Various Subclasses: 1=Axe, 5=Mace, 6=Polearm, 8=Sword, 17=Knife/Dagger sometimes, etc)
        # Add your specific 2H subclasses here
        if sub in [1, 5, 6, 8, 17]: return 1
        
        # Everything else (1H swords, daggers, maces)
        return 3
        
    return 0 # Default fallback
def parse_ilvl_input(x):
    if '-' in x:
        low, high = map(int, x.split('-'))
        if not (10 <= low <= 284 and 10 <= high <= 284 and low <= high): raise Exception
        return (low, high)
    else:
        val = int(x)
        if not (10 <= val <= 284): raise Exception
        return (val, val)
def get_db_connection():
    return mysql.connector.connect(**DB_CONFIG)

def get_appropriate_req_level(cursor, ilvl, quality):
    # Search for items within +/- 3 levels of target ilvl, same quality
    query = """
    SELECT AVG(RequiredLevel) 
    FROM item_template 
    WHERE itemlevel BETWEEN %s AND %s 
    AND Quality = %s 
    AND RequiredLevel > 1
    """
    cursor.execute(query, (ilvl - 10, ilvl + 10, quality))
    result = cursor.fetchone()[0]
    
    # Calculate initial value
    req_level = int(result) if result else max(1, int(ilvl * 0.75))
    
    # --- LEVEL CLAMPING LOGIC ---
    # 1. Cap for iLevel 90 and below (Raiding bracket)
    if ilvl <= 90:
        req_level = min(req_level, 60)
        
    # 2. Cap for iLevel 154 and below (BC Raiding bracket)
    elif ilvl <= 154:
        req_level = min(req_level, 70)
        
    # 3. Global absolute cap
    req_level = min(req_level, 80)
    
    return req_level
    
def interpolate_armor(nodes, target_ilvl):
    valid_nodes = [n for n in nodes if n["avg_armor"] > 0]
    if not valid_nodes: return 0.0
    for n in valid_nodes:
        if n["itemlevel"] == target_ilvl: return n["avg_armor"]
    low_node  = max([n for n in valid_nodes if n["itemlevel"] <= target_ilvl], key=lambda x: x["itemlevel"], default=None)
    high_node = min([n for n in valid_nodes if n["itemlevel"] >= target_ilvl], key=lambda x: x["itemlevel"], default=None)
    if low_node and high_node and low_node != high_node:
        x0, x1 = low_node["itemlevel"], high_node["itemlevel"]
        t = (target_ilvl - x0) / (x1 - x0)
        return low_node["avg_armor"] + t * (high_node["avg_armor"] - low_node["avg_armor"])
    # Clamp to nearest endpoint rather than falling back to valid_nodes[0]
    # (which is the LOWEST ilvl node, not the closest one).
    # This prevents a shield at ilvl 134 from inheriting ilvl 10 armor values
    # just because the curve has no data above ilvl 115.
    if low_node:  return low_node["avg_armor"]
    if high_node: return high_node["avg_armor"]
    return valid_nodes[-1]["avg_armor"]

def interpolate_block(nodes, target_ilvl):
    """Identical structure to interpolate_armor but reads avg_block.
    Returns 0.0 for non-shields which have no block data in their ilvl_sheet.
    Same endpoint-clamping fix applied so high-ilvl shields never get ilvl-10 block.
    """
    valid_nodes = [n for n in nodes if n.get("avg_block", 0) > 0]
    if not valid_nodes: return 0.0
    for n in valid_nodes:
        if n["itemlevel"] == target_ilvl: return n["avg_block"]
    low_node  = max([n for n in valid_nodes if n["itemlevel"] <= target_ilvl], key=lambda x: x["itemlevel"], default=None)
    high_node = min([n for n in valid_nodes if n["itemlevel"] >= target_ilvl], key=lambda x: x["itemlevel"], default=None)
    if low_node and high_node and low_node != high_node:
        x0, x1 = low_node["itemlevel"], high_node["itemlevel"]
        t = (target_ilvl - x0) / (x1 - x0)
        return low_node["avg_block"] + t * (high_node["avg_block"] - low_node["avg_block"])
    if low_node:  return low_node["avg_block"]
    if high_node: return high_node["avg_block"]
    return valid_nodes[-1]["avg_block"]


# ---------------------------------------------------------------------------
# Cross-quality armor / block scaling
# ---------------------------------------------------------------------------
# Problem: when a high-quality (e.g. Epic) item has no reference items at or
# below the requested ilvl, interpolate_armor/block clamps to the nearest node
# ABOVE the target, which can be drastically overpowered.
# Example: ilvl 24 Epic shield -- nearest Epic node is Green Tower at ilvl 41
# (1507 armor). Clamping to that gives a level-24 shield 3x the correct armor.
#
# Solution (applied only when this extrapolation gap is detected):
#   1. Find the nearest reference point in the Epic curve (the "anchor" ilvl).
#   2. Look up a lower-quality (Rare → Uncommon) curve at that same anchor ilvl.
#   3. Compute the ratio: epic_armor_at_anchor / lowerqual_armor_at_anchor.
#      This is the "quality premium" for that slot at that ilvl.
#   4. Look up the lower-quality armor at the actual TARGET ilvl (where data
#      almost always exists because green/blue items are far more numerous).
#   5. Apply the quality premium to get the estimated epic value.
#
# Quality fallback order: 4 (Epic) → 3 (Rare) → 2 (Uncommon)
# Tolerance for "anchor ilvl" neighbor search: ±5 ilvl.
# This path is only taken when the target quality sheet has NO node at or
# below the target ilvl (i.e. the only reference is above the target).

_QUALITY_FALLBACK_ORDER = {4: [3, 2], 3: [2], 2: []}  # qual -> lower quals to try
_ANCHOR_TOLERANCE = 5   # how many ilvls away the anchor node is allowed to be


def _needs_cross_quality_scaling(nodes, field, target_ilvl):
    """Return True when the only available data is ABOVE target_ilvl."""
    valid = [n for n in nodes if n.get(field, 0) > 0]
    if not valid:
        return False  # no data at all; let the caller handle it
    has_low = any(n["itemlevel"] <= target_ilvl for n in valid)
    return not has_low  # True only when every node is strictly above target


def _cross_quality_scale(source_dict, cls, subcls, inv_type, qual,
                         target_ilvl, field, interp_fn):
    """
    Apply cross-quality ratio scaling when the requested quality has no
    reference nodes at or below target_ilvl.  Works for any field:
    avg_armor, avg_block, avg_budget, avg_dps.

    source_dict  -- the dict to look up lower-quality sheets from
                    (slot_budget_curves for budget, lookup_database for others)
    interp_fn    -- the raw interpolation function for this field

    Returns the scaled value, or None if cross-quality path also fails.
    """
    own_sheet  = source_dict.get((cls, subcls, inv_type, qual), [])
    own_valid  = [n for n in own_sheet if n.get(field, 0) > 0]
    if not own_valid:
        return None

    # Nearest node in the current quality sheet (will be above target)
    anchor_node = min(own_valid, key=lambda n: abs(n["itemlevel"] - target_ilvl))
    anchor_ilvl = anchor_node["itemlevel"]
    anchor_val  = anchor_node[field]

    for lower_qual in _QUALITY_FALLBACK_ORDER.get(qual, []):
        lower_sheet = source_dict.get((cls, subcls, inv_type, lower_qual), [])
        lower_valid = [n for n in lower_sheet if n.get(field, 0) > 0]
        if not lower_valid:
            continue

        lq_anchor = min(lower_valid, key=lambda n: abs(n["itemlevel"] - anchor_ilvl))
        if abs(lq_anchor["itemlevel"] - anchor_ilvl) > _ANCHOR_TOLERANCE:
            continue
        lq_anchor_val = lq_anchor[field]
        if lq_anchor_val <= 0:
            continue

        ratio = anchor_val / lq_anchor_val

        lq_at_target = interp_fn(lower_sheet, target_ilvl)
        if lq_at_target <= 0:
            continue

        return lq_at_target * ratio

    return None


def get_scaled_armor(lookup_database, category, quality_code, ilvl, sheet):
    """
    Drop-in replacement for interpolate_armor at call sites.
    Uses cross-quality scaling when the target quality has no lower-ilvl nodes.
    """
    cls, subcls, inv_type = category["class"], category["subclass"], category["InventoryType"]
    if _needs_cross_quality_scaling(sheet, "avg_armor", ilvl):
        scaled = _cross_quality_scale(
            lookup_database, cls, subcls, inv_type, quality_code,
            ilvl, "avg_armor", interpolate_armor
        )
        if scaled is not None:
            return scaled
    return interpolate_armor(sheet, ilvl)


def get_scaled_block(lookup_database, category, quality_code, ilvl, sheet):
    """
    Drop-in replacement for interpolate_block at call sites.
    Uses cross-quality scaling when the target quality has no lower-ilvl nodes.
    """
    cls, subcls, inv_type = category["class"], category["subclass"], category["InventoryType"]
    if _needs_cross_quality_scaling(sheet, "avg_block", ilvl):
        scaled = _cross_quality_scale(
            lookup_database, cls, subcls, inv_type, quality_code,
            ilvl, "avg_block", interpolate_block
        )
        if scaled is not None:
            return scaled
    return interpolate_block(sheet, ilvl)


def get_next_entry_id(cursor):
    cursor.execute(f"SELECT MAX(entry) FROM item_template WHERE entry >= {ENTRY_ID_START}")
    result = cursor.fetchone()[0]
    return (result + 1) if result else ENTRY_ID_START
def generate_item_name(category_config):
    """
    Pure operational execution function. Uses the global 'master' arrays
    unpacked at the top of the file to roll accurate Blizzlike names.
    """
    # 1. Parse configuration inputs (handles dict, 2-tuples, or 3-tuples)
    if isinstance(category_config, dict):
        cls = category_config.get("class")
        subcls = category_config.get("subclass")
        inv_type = category_config.get("InventoryType")
    elif isinstance(category_config, (tuple, list)):
        cls = category_config[0]
        subcls = category_config[1]
        inv_type = category_config[2] if len(category_config) >= 3 else None
    else:
        cls, subcls, inv_type = 2, 0, None

    cat_keys = (cls, subcls)
    
    # 2. Extract context lists from the global name_database
    ndb = name_database.get(cat_keys, {
        "adjectives": ["Reinforced"], "materials": ["Iron"], "nouns": ["Blade"], "properties": ["of Power"]
    })
    
    adj_pool = ndb.get("adjectives", ["Reinforced"])
    mat_pool = ndb.get("materials", ["Iron"])
    prop_pool = ndb.get("properties", ["of Power"])
    
    # 3. Pull Curated Noun Seeds from the unpacked global blocks
    noun_pool = ndb.get("nouns", ["Blade"])
    
    if cls == 2:  # Weapons Mapping
        noun_pool = weapon_nouns.get(subcls, noun_pool)
        
    elif cls == 4 and inv_type is not None:  # Armor & Accessories Mapping
        if inv_type in armor_nouns:
            slot_map = armor_nouns[inv_type]
            # Match exact material style subclass, fall back to slot standard (0), or pick first available
            noun_pool = slot_map.get(subcls) or slot_map.get(0) or list(slot_map.values())[0]

    # 4. Generate & Sanitize final string output
    loops = 0
    while True:
        noun = random.choice(noun_pool) if noun_pool else "Item"
        adj = random.choice(adj_pool) if (adj_pool and random.random() < 0.6) else ""
        mat = random.choice(mat_pool) if (mat_pool and random.random() < 0.4) else ""
        prop = random.choice(prop_pool) if (prop_pool and random.random() < 0.3) else ""
        
        final_name = " ".join(f"{adj} {mat} {noun} {prop}".split())
        
        loops += 1
        if len(final_name.split()) >= 2 or loops > 20:
            return final_name

def get_appropriate_display_id(cat_keys, target_ilvl, target_qual, target_inv_type): # Add argument
    ndb = name_database.get(cat_keys)
    if not ndb or "displays" not in ndb: return {"id": 42531, "lvl": 0, "q": 0}
    
    pool = ndb["displays"]
    valid_qualities = [target_qual - 1, target_qual, target_qual + 1]
    min_lvl, max_lvl = target_ilvl * 0.8, target_ilvl * 1.2
    
    candidates = [
        d for d in pool 
        if d["q"] in valid_qualities 
        and min_lvl <= d["lvl"] <= max_lvl 
        and d.get("InventoryType") == target_inv_type
    ]
    
    return random.choice(candidates) if candidates else random.choice(pool)
    
    # Return match or random fallback from the category pool
    return random.choice(candidates) if candidates else random.choice([d["id"] for d in pool])
# Smoothed-curve cache.  Keys are either the 4-tuple (cls, subcls, inv_type, qual)
# for per-slot curves or the 2-tuple (inv_type, qual) for the coarser fallback.
_smoothed_budget_cache = {}

def _smooth_nodes(raw_nodes):
    """Monotonic forward-pass smoother (isotonic regression, 1-D).
    Ensures avg_budget never decreases as itemlevel increases."""
    if not raw_nodes:
        return []
    sorted_nodes = sorted(raw_nodes, key=lambda n: n["itemlevel"])
    smoothed, running_max = [], 0.0
    for n in sorted_nodes:
        clamped = max(n["avg_budget"], running_max)
        running_max = clamped
        smoothed.append({"itemlevel": n["itemlevel"], "avg_budget": clamped})
    return smoothed

def _get_smoothed_nodes(cls, subcls, inv_type, qual):
    """
    Returns monotonically-smoothed budget nodes for a specific slot.
    Lookup priority:
      1. slot_budget_curves[(cls, subcls, inv_type, qual)]  -- most specific
      2. global_budget_curves[(inv_type, qual)]             -- coarse fallback
    Both are smoothed on first access and cached.
    """
    cache_key = (cls, subcls, inv_type, qual)
    if cache_key in _smoothed_budget_cache:
        return _smoothed_budget_cache[cache_key]

    # Try fine-grained per-slot curve first
    raw = slot_budget_curves.get(cache_key)
    if not raw:
        # Fall back to coarse (inv_type, qual) curve
        raw = global_budget_curves.get((inv_type, qual), [])

    result = _smooth_nodes(raw)
    _smoothed_budget_cache[cache_key] = result
    return result

def interpolate_macro_budget(cls, subcls, inv_type, qual, target_ilvl):
    nodes = _get_smoothed_nodes(cls, subcls, inv_type, qual)
    if not nodes: return 0.0
    if len(nodes) == 1: return nodes[0]["avg_budget"]

    # Exact match
    for n in nodes:
        if n["itemlevel"] == target_ilvl: return n["avg_budget"]

    # Linear interpolation between bracketing nodes
    low_node, high_node = None, None
    for n in nodes:
        if n["itemlevel"] < target_ilvl:
            low_node = n
        elif n["itemlevel"] > target_ilvl and high_node is None:
            high_node = n
            break

    if low_node is not None and high_node is not None:
        x0, x1 = low_node["itemlevel"], high_node["itemlevel"]
        t = (target_ilvl - x0) / (x1 - x0)
        return low_node["avg_budget"] + t * (high_node["avg_budget"] - low_node["avg_budget"])

    # Only high_node exists: no reference at or below target_ilvl.
    # Apply cross-quality ratio scaling instead of clamping to the high node,
    # which would give a low-ilvl item the budget of a much higher-ilvl item.
    if low_node is None and high_node is not None:
        # Build a thin wrapper so _cross_quality_scale can call interp_fn(sheet, ilvl)
        def _interp_budget(budget_nodes, tgt):
            # budget_nodes are {itemlevel, avg_budget} dicts from slot_budget_curves
            vn = [n for n in budget_nodes if n.get("avg_budget", 0) > 0]
            if not vn: return 0.0
            for n in vn:
                if n["itemlevel"] == tgt: return n["avg_budget"]
            lo = max([n for n in vn if n["itemlevel"] <= tgt], key=lambda x: x["itemlevel"], default=None)
            hi = min([n for n in vn if n["itemlevel"] >= tgt], key=lambda x: x["itemlevel"], default=None)
            if lo and hi and lo != hi:
                t = (tgt - lo["itemlevel"]) / (hi["itemlevel"] - lo["itemlevel"])
                return lo["avg_budget"] + t * (hi["avg_budget"] - lo["avg_budget"])
            if lo: return lo["avg_budget"]
            if hi: return hi["avg_budget"]
            return vn[-1]["avg_budget"]

        scaled = _cross_quality_scale(
            slot_budget_curves, cls, subcls, inv_type, qual,
            target_ilvl, "avg_budget", _interp_budget
        )
        if scaled is None:
            # slot_budget_curves failed; try global_budget_curves as source
            def _global_source_dict():
                # Wrap global_budget_curves into the (cls,subcls,inv,qual) key format
                # so _cross_quality_scale can look up lower qualities
                wrapped = {}
                for (it, q), v in global_budget_curves.items():
                    wrapped[(cls, subcls, it, q)] = v
                return wrapped
            scaled = _cross_quality_scale(
                _global_source_dict(), cls, subcls, inv_type, qual,
                target_ilvl, "avg_budget", _interp_budget
            )
        if scaled is not None:
            return scaled
        # Final fallback: clamp to high_node (old behaviour, now last resort)
        return high_node["avg_budget"]

    if low_node is not None: return low_node["avg_budget"]
    return nodes[0]["avg_budget"]

def interpolate_local_dps(nodes, target_ilvl, lookup_database=None,
                          cls=None, subcls=None, inv_type=None, qual=None):
    """Interpolate avg_dps for weapons.  When no lower-ilvl node exists and
    caller supplies lookup_database + slot keys, applies cross-quality scaling
    rather than clamping to the nearest (too-high) reference node."""
    valid_nodes = [n for n in nodes if n["avg_dps"] > 0]
    if not valid_nodes: return 0.0
    for n in valid_nodes:
        if n["itemlevel"] == target_ilvl: return n["avg_dps"]
    if len(valid_nodes) == 1: return valid_nodes[0]["avg_dps"]

    low_node  = max([n for n in valid_nodes if n["itemlevel"] <= target_ilvl], key=lambda x: x["itemlevel"], default=None)
    high_node = min([n for n in valid_nodes if n["itemlevel"] >= target_ilvl], key=lambda x: x["itemlevel"], default=None)

    if low_node and high_node and low_node != high_node:
        t = (target_ilvl - low_node["itemlevel"]) / (high_node["itemlevel"] - low_node["itemlevel"])
        return low_node["avg_dps"] + t * (high_node["avg_dps"] - low_node["avg_dps"])

    # No lower node: apply cross-quality scaling if possible
    if low_node is None and lookup_database is not None and cls is not None:
        def _interp_dps(dps_nodes, tgt):
            vn = [n for n in dps_nodes if n.get("avg_dps", 0) > 0]
            if not vn: return 0.0
            for n in vn:
                if n["itemlevel"] == tgt: return n["avg_dps"]
            lo = max([n for n in vn if n["itemlevel"] <= tgt], key=lambda x: x["itemlevel"], default=None)
            hi = min([n for n in vn if n["itemlevel"] >= tgt], key=lambda x: x["itemlevel"], default=None)
            if lo and hi and lo != hi:
                t = (tgt - lo["itemlevel"]) / (hi["itemlevel"] - lo["itemlevel"])
                return lo["avg_dps"] + t * (hi["avg_dps"] - lo["avg_dps"])
            if lo: return lo["avg_dps"]
            if hi: return hi["avg_dps"]
            return vn[-1]["avg_dps"]

        scaled = _cross_quality_scale(
            lookup_database, cls, subcls, inv_type, qual,
            target_ilvl, "avg_dps", _interp_dps
        )
        if scaled is not None:
            return scaled

    if low_node:  return low_node["avg_dps"]
    if high_node: return high_node["avg_dps"]
    return valid_nodes[0]["avg_dps"]

def get_interpolated_properties(ilvl_sheet, target_ilvl, cls, subcls, inv_type, qual):
    if not ilvl_sheet: return None
    final_budget = interpolate_macro_budget(cls, subcls, inv_type, qual, target_ilvl)
    # Pass slot keys so interpolate_local_dps can apply cross-quality DPS scaling
    final_dps = interpolate_local_dps(
        ilvl_sheet, target_ilvl,
        lookup_database=lookup_database,
        cls=cls, subcls=subcls, inv_type=inv_type, qual=qual
    )
    closest_node = min(ilvl_sheet, key=lambda x: abs(x["itemlevel"] - target_ilvl))
    return {
        "itemlevel": target_ilvl, "avg_budget": final_budget, "avg_dps": final_dps,
        "display_ids": closest_node["display_ids"] if closest_node["display_ids"] else [42531],
        "stat_profiles": closest_node["stat_profiles"]
    }

internal_memory = []
generated_itemsets = []  # each: {"id": int, "name": str, "item_indices": [...], "spells": [(threshold, spellid), ...]}

def prompt_blueprint(label, options_map):
    """
    Displays a blueprint selection menu and returns (chosen_blueprint_key, candidates).
    options_map: dict of display_number -> blueprint_key string, e.g. {1: "AGI_DPS", 2: "STR_DPS"}
    Returns:
      chosen_blueprint_key  – the selected key, or "__ALL__" for random-per-item
      chosen_blueprint_candidates – list of all keys in options_map (used when __ALL__)
    """
    all_num = max(options_map) + 1
    print(f"\nSelect {label} Archetype Variant:")
    for num, key in options_map.items():
        # Pretty-print the blueprint key
        friendly = key.replace("_", " ").title()
        print(f"  [{num}] {friendly}")
    print(f"  [{all_num}] All of the above (random per item)")
    valid = list(options_map.keys()) + [all_num]
    bp_choice = get_input("Choice: ", lambda x: int(x) if int(x) in valid else int("err"))
    candidates = list(options_map.values())
    if bp_choice == all_num:
        return "__ALL__", candidates
    return options_map[bp_choice], candidates

def get_input(prompt, validation_fn):
    while True:
        try:
            val = input(prompt).strip()
            return validation_fn(val)
        except Exception:
            print(f"⚠️ Invalid choice. Attempt verification again.")

print("=======================================================")
print("  🔮 COHESIVE SYNERGY ARCHETYPE GENERATOR ENGINE       ")
print("=======================================================")

# ── Helper: resolve a list of CATEGORIES keys from a menu that supports "All of the above" ──

def pick_cat_indices(prompt_label, mapping):
    """
    Shows a numbered menu from `mapping` (int -> cat_idx or list[cat_idx])
    plus an extra "All of the above" option.
    Returns a flat list of cat_idx values chosen.
    mapping values can be a single int or a list of ints.
    """
    all_num = max(mapping.keys()) + 1
    print(f"\n{prompt_label}")
    for num, val in mapping.items():
        if isinstance(val, list):
            label = " / ".join(CATEGORIES[v]["name"] for v in val)
        else:
            label = CATEGORIES[val]["name"]
        print(f"  [{num}] {label}")
    print(f"  [{all_num}] All of the above (random per item)")
    valid = list(mapping.keys()) + [all_num]
    choice = get_input("Select: ", lambda x: int(x) if int(x) in valid else int("err"))
    if choice == all_num:
        # flatten all values
        result = []
        for v in mapping.values():
            if isinstance(v, list):
                result.extend(v)
            else:
                result.append(v)
        return result
    else:
        val = mapping[choice]
        return val if isinstance(val, list) else [val]


def resolve_blueprint_for_category(category):
    """
    Given a single CATEGORIES entry (dict), prompt for and return
    (chosen_blueprint_key, chosen_blueprint_candidates).
    """
    subc = category["subclass"]
    cls  = category["class"]

    if cls == 2:  # WEAPONS
        if subc == 19:   # Wand
            return prompt_blueprint("Wand", {1: "SP_DPS", 2: "HEALER"})
        elif subc == 7:  # 1H Sword
            return prompt_blueprint("1H Sword", {1: "AGI_DPS", 2: "STR_TANK", 3: "STR_DPS", 4: "SP_DPS", 5: "HEALER"})
        elif subc == 8:  # 2H Sword
            return prompt_blueprint("2H Sword", {1: "AGI_DPS", 2: "STR_DPS", 3: "STR_TANK", 4: "AGI_INT_DPS"})
        elif subc == 15: # Dagger
            return prompt_blueprint("Dagger", {1: "AGI_DPS", 2: "SP_DPS", 3: "HEALER", 4: "AGI_INT_DPS"})
        elif subc == 10: # Staff
            return prompt_blueprint("Staff", {1: "AGI_DPS", 2: "SP_DPS", 3: "HEALER", 4: "AGI_TANK", 5: "STR_DPS", 6: "AGI_INT_DPS"})
        elif subc == 13: # Fist Weapon
            return prompt_blueprint("Fist Weapon", {1: "AGI_DPS", 2: "SP_DPS", 3: "AGI_INT_DPS"})
        elif subc == 0:  # 1H Axe
            return prompt_blueprint("1H Axe", {1: "AGI_DPS", 2: "STR_DPS", 3: "STR_TANK", 4: "AGI_INT_DPS"})
        elif subc == 1:  # 2H Axe
            return prompt_blueprint("2H Axe", {1: "AGI_DPS", 2: "STR_DPS", 3: "STR_TANK", 4: "AGI_INT_DPS"})
        elif subc in [5, 6]:  # 2H Mace or Polearm
            return prompt_blueprint(category['name'], {1: "STR_DPS", 2: "AGI_DPS", 3: "STR_TANK", 4: "AGI_TANK", 5: "AGI_INT_DPS"})
        elif subc in [2, 3, 18]:  # Bows, Guns, Crossbows
            return prompt_blueprint(category['name'], {1: "STR_TANK", 2: "STR_DPS", 3: "AGI_DPS", 4: "AGI_INT_DPS"})
        elif subc == 16: # Thrown
            return prompt_blueprint("Thrown Weapon", {1: "AGI_DPS", 2: "STR_DPS", 3: "STR_TANK"})
        elif subc == 4:  # 1H Mace
            return prompt_blueprint("1H Mace", {1: "STR_DPS", 2: "STR_TANK", 3: "HEALER", 4: "SP_DPS"})

    elif cls == 4:  # ARMOR
        if subc == 1:    # Cloth
            # Intercept Cloaks (InventoryType 16) to use Miscellaneous blueprints
            if category.get('InventoryType') == 16:
                return prompt_blueprint("Cloak", 
                    {1: "AGI_DPS", 2: "STR_DPS", 3: "SP_DPS", 4: "HEALER", 
                     5: "AGI_INT_DPS", 6: "STR_TANK", 7: "AGI_TANK"})
            # Standard Cloth logic
            return prompt_blueprint("Cloth", {1: "SP_DPS", 2: "HEALER"})
        elif subc == 2:  # Leather
            return prompt_blueprint("Leather", {1: "AGI_DPS", 2: "AGI_TANK", 3: "HEALER", 4: "SP_DPS", 5: "AGI_INT_DPS"})
        elif subc == 3:  # Mail
            return prompt_blueprint("Mail", {1: "AGI_DPS", 2: "STR_DPS", 3: "HEALER", 4: "SP_DPS", 5: "AGI_INT_DPS"})
        elif subc == 4:  # Plate
            return prompt_blueprint("Plate", {1: "STR_DPS", 2: "STR_TANK", 3: "HEALER"})
        elif subc == 0 or subc == 6:  # Miscellaneous or Shield
            if subc == 6:
                return prompt_blueprint(category['name'], {1: "STR_TANK", 2: "SP_DPS", 3: "HEALER"})
            elif category['InventoryType'] == 23:
                return prompt_blueprint(category['name'], {1: "SP_DPS", 2: "HEALER"})
            else:
                return prompt_blueprint(category['name'],
                    {1: "AGI_DPS", 2: "STR_DPS", 3: "SP_DPS", 4: "HEALER",
                     5: "AGI_INT_DPS", 6: "STR_TANK", 7: "AGI_TANK"})
        else:
            print(f"⚠️ Config missing for Subclass {subc} (Class {cls}). Defaulting to STR_DPS.")
            return "STR_DPS", ["STR_DPS"]

    # Fallback
    return "STR_DPS", ["STR_DPS"]


# ── MASS CREATION: blueprint auto-resolver (no user prompts) ─────────────────
# Maps each category's (class, subclass, InventoryType) signature to the full
# list of valid blueprints so mass-creation can pick one at random per item.

MASS_BLUEPRINT_MAP = {
    # Weapons ─────────────────────────────────────────────────────────────────
    (2, 19, 26): ["SP_DPS", "HEALER"],                            # Wand
    (2,  7, 13): ["AGI_DPS", "STR_TANK", "STR_DPS", "SP_DPS", "HEALER"],  # 1H Sword
    (2,  8, 17): ["AGI_DPS", "STR_DPS", "STR_TANK", "AGI_INT_DPS"],       # 2H Sword
    (2, 15, 13): ["AGI_DPS", "SP_DPS", "HEALER", "AGI_INT_DPS"],          # Dagger
    (2, 10, 17): ["AGI_DPS", "SP_DPS", "HEALER", "AGI_TANK", "STR_DPS", "AGI_INT_DPS"],  # Staff
    (2, 13, 13): ["AGI_DPS", "SP_DPS", "AGI_INT_DPS"],           # Fist Weapon
    (2,  0, 13): ["AGI_DPS", "STR_DPS", "STR_TANK", "AGI_INT_DPS"],  # 1H Axe
    (2,  1, 17): ["AGI_DPS", "STR_DPS", "STR_TANK", "AGI_INT_DPS"],  # 2H Axe
    (2,  5, 17): ["STR_DPS", "AGI_DPS", "STR_TANK", "AGI_TANK", "AGI_INT_DPS"],  # 2H Mace
    (2,  6, 17): ["STR_DPS", "AGI_DPS", "STR_TANK", "AGI_TANK", "AGI_INT_DPS"],  # Polearm
    (2,  2, 15): ["STR_TANK", "STR_DPS", "AGI_DPS", "AGI_INT_DPS"],  # Bow
    (2, 18, 26): ["STR_TANK", "STR_DPS", "AGI_DPS", "AGI_INT_DPS"],  # Crossbow
    (2,  3, 26): ["STR_TANK", "STR_DPS", "AGI_DPS", "AGI_INT_DPS"],  # Gun
    (2, 16, 25): ["AGI_DPS", "STR_DPS", "STR_TANK"],              # Thrown
    (2,  4, 13): ["STR_DPS", "STR_TANK", "HEALER", "SP_DPS"],    # 1H Mace
    # Armor – Cloth ───────────────────────────────────────────────────────────
    (4,  1,  1): ["SP_DPS", "HEALER"],   # Cloth Helm
    (4,  1,  3): ["SP_DPS", "HEALER"],   # Cloth Shoulders
    (4,  1,  5): ["SP_DPS", "HEALER"],   # Cloth Chest
    (4,  1,  9): ["SP_DPS", "HEALER"],   # Cloth Wrist
    (4,  1, 10): ["SP_DPS", "HEALER"],   # Cloth Gloves
    (4,  1,  6): ["SP_DPS", "HEALER"],   # Cloth Waist
    (4,  1,  7): ["SP_DPS", "HEALER"],   # Cloth Legs
    (4,  1,  8): ["SP_DPS", "HEALER"],   # Cloth Feet
    # Armor – Leather ─────────────────────────────────────────────────────────
    (4,  2,  1): ["AGI_DPS", "AGI_TANK", "HEALER", "SP_DPS", "AGI_INT_DPS"],
    (4,  2,  3): ["AGI_DPS", "AGI_TANK", "HEALER", "SP_DPS", "AGI_INT_DPS"],
    (4,  2,  5): ["AGI_DPS", "AGI_TANK", "HEALER", "SP_DPS", "AGI_INT_DPS"],
    (4,  2,  9): ["AGI_DPS", "AGI_TANK", "HEALER", "SP_DPS", "AGI_INT_DPS"],
    (4,  2, 10): ["AGI_DPS", "AGI_TANK", "HEALER", "SP_DPS", "AGI_INT_DPS"],
    (4,  2,  6): ["AGI_DPS", "AGI_TANK", "HEALER", "SP_DPS", "AGI_INT_DPS"],
    (4,  2,  7): ["AGI_DPS", "AGI_TANK", "HEALER", "SP_DPS", "AGI_INT_DPS"],
    (4,  2,  8): ["AGI_DPS", "AGI_TANK", "HEALER", "SP_DPS", "AGI_INT_DPS"],
    # Armor – Mail ────────────────────────────────────────────────────────────
    (4,  3,  1): ["AGI_DPS", "STR_DPS", "STR_TANK", "HEALER", "SP_DPS", "AGI_INT_DPS"],
    (4,  3,  3): ["AGI_DPS", "STR_DPS", "STR_TANK", "HEALER", "SP_DPS", "AGI_INT_DPS"],
    (4,  3,  5): ["AGI_DPS", "STR_DPS", "STR_TANK", "HEALER", "SP_DPS", "AGI_INT_DPS"],
    (4,  3,  9): ["AGI_DPS", "STR_DPS", "STR_TANK", "HEALER", "SP_DPS", "AGI_INT_DPS"],
    (4,  3, 10): ["AGI_DPS", "STR_DPS", "STR_TANK", "HEALER", "SP_DPS", "AGI_INT_DPS"],
    (4,  3,  6): ["AGI_DPS", "STR_DPS", "STR_TANK", "HEALER", "SP_DPS", "AGI_INT_DPS"],
    (4,  3,  7): ["AGI_DPS", "STR_DPS", "STR_TANK", "HEALER", "SP_DPS", "AGI_INT_DPS"],
    (4,  3,  8): ["AGI_DPS", "STR_DPS", "STR_TANK", "HEALER", "SP_DPS", "AGI_INT_DPS"],
    # Armor – Plate ───────────────────────────────────────────────────────────
    (4,  4,  1): ["STR_DPS", "STR_TANK", "HEALER"],
    (4,  4,  3): ["STR_DPS", "STR_TANK", "HEALER"],
    (4,  4,  5): ["STR_DPS", "STR_TANK", "HEALER"],
    (4,  4,  9): ["STR_DPS", "STR_TANK", "HEALER"],
    (4,  4, 10): ["STR_DPS", "STR_TANK", "HEALER"],
    (4,  4,  6): ["STR_DPS", "STR_TANK", "HEALER"],
    (4,  4,  7): ["STR_DPS", "STR_TANK", "HEALER"],
    (4,  4,  8): ["STR_DPS", "STR_TANK", "HEALER"],
    # Misc / Accessories ──────────────────────────────────────────────────────
    (4,  1, 16): ["AGI_DPS", "STR_DPS", "SP_DPS", "HEALER", "AGI_INT_DPS", "STR_TANK", "AGI_TANK"],  # Cloak
    (4,  0,  2): ["AGI_DPS", "STR_DPS", "SP_DPS", "HEALER", "AGI_INT_DPS", "STR_TANK", "AGI_TANK"],  # Necklace
    (4,  0, 11): ["AGI_DPS", "STR_DPS", "SP_DPS", "HEALER", "AGI_INT_DPS", "STR_TANK", "AGI_TANK"],  # Ring
    (4,  6, 14): ["STR_TANK", "SP_DPS", "HEALER"],               # Shield
    (4,  0, 23): ["SP_DPS", "HEALER"],                            # Offhand
}

def get_mass_blueprint(category):
    """Return a random valid blueprint key for a category without any user prompts."""
    sig = (category["class"], category["subclass"], category["InventoryType"])
    pool = MASS_BLUEPRINT_MAP.get(sig)
    if pool:
        return random.choice(pool)
    # Generic fallback by material/class
    cls, subc = category["class"], category["subclass"]
    if cls == 2:
        return random.choice(["AGI_DPS", "STR_DPS", "SP_DPS"])
    elif subc == 4:   # Plate
        return random.choice(["STR_DPS", "STR_TANK", "HEALER"])
    elif subc == 1:   # Cloth
        return random.choice(["SP_DPS", "HEALER"])
    else:
        return random.choice(list(BLUEPRINTS.keys()))


def run_mass_creation():
    """
    Mass Creation mode: skips all archetype prompts.
    Asks only: scope (Weapons/Armor/All), Quality, ilvl range,
    Budget Deviation, Stat Split Deviation, Stat Density, and Item Count.
    Then generates items using a smart anti-repetition scheduler that
    spreads types and levels as evenly as possible before repeating.
    """
    print("\n" + "="*57)
    print("  ⚡ MASS CREATION ENGINE — Automated Population Mode   ")
    print("="*57)

    # ── Scope ────────────────────────────────────────────────────────────────
    print("\nSelect Item Scope:")
    print("  [1] Weapons only")
    print("  [2] Armor only")
    print("  [3] All (Weapons + Armor)")
    scope = get_input("Choice: ", lambda x: int(x) if x in ['1', '2', '3'] else int("err"))

    if scope == 1:
        cat_indices = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15]
    elif scope == 2:
        cat_indices = list(range(16, 53))
    else:
        cat_indices = list(CATEGORIES.keys())

    # ── Quality ──────────────────────────────────────────────────────────────
    print("\nAvailable Qualities:")
    for k, v in QUALITIES.items():
        print(f"  [{k}] {v['name']}")
    selected_quality_entry = QUALITIES[get_input(
        "Select Item Quality (Number): ",
        lambda x: int(x) if int(x) in QUALITIES else int("err")
    )]
    quality_code_pool = selected_quality_entry["code"] if selected_quality_entry["multi"] else [selected_quality_entry["code"]]

    # ── ilvl range ───────────────────────────────────────────────────────────
    print("\nEnter Target Item Level range (e.g. '10-284' for full game, or '50-80'):")
    ilvl_range = get_input("Choice: ", parse_ilvl_input)

    # ── Budget Deviation ─────────────────────────────────────────────────────
    variance = get_input(
        "Enter Budget Quality Variance % (0 to 25, recommended 8): ",
        lambda x: float(x) / 100.0 if 0 <= float(x) <= 25 else float("err")
    )

    # ── Stat Split Deviation ─────────────────────────────────────────────────
    print("\nStat Split Deviation — controls how unevenly stats are spread across slots.")
    skew_factor = get_input(
        "Enter Max Stat Split Deviation % (0 = even, 40 = recommended variance): ",
        lambda x: float(x) if 0 <= float(x) <= 100 else float("err")
    )
    dist_mode = 2 if skew_factor > 0 else 1

    # ── Stat Density ─────────────────────────────────────────────────────────
    print("\nStat Density Rule:")
    print("  [1] Database-Driven (most authentic)")
    print("  [2] Progressive Blizzlike (scales with ilvl)")
    print("  [3] Explicit Manual Count")
    print("  [4] Random Range (rolls a random count per item)")
    density_mode = get_input("Select (1/2/3/4): ", lambda x: int(x) if int(x) in [1,2,3,4] else int("err"))
    chosen_density_count = get_input(
        "Enter exact stat count (1-6): ",
        lambda x: int(x) if 1 <= int(x) <= 6 else int("err")
    ) if density_mode == 3 else 0
    density_range_min, density_range_max = 0, 0
    if density_mode == 4:
        density_range_min = get_input(
            "Enter minimum stat count (1-6): ",
            lambda x: int(x) if 1 <= int(x) <= 6 else int("err")
        )
        density_range_max = get_input(
            f"Enter maximum stat count ({density_range_min}-6): ",
            lambda x: int(x) if density_range_min <= int(x) <= 6 else int("err")
        )

    # ── Item Binding ─────────────────────────────────────────────────────────
    print("\nItem Binding:")
    print("  [1] Bind on Equip (BoE)")
    print("  [2] Bind on Pickup (BoP)")
    print("  [3] Both (mixed)")
    bonding_mode = get_input("Choice: ", lambda x: int(x) if int(x) in [1, 2, 3] else int("err"))
    boe_ratio = 0.5
    if bonding_mode == 3:
        boe_ratio = get_input(
            "Proportion of BoE items (0.0 = all BoP, 1.0 = all BoE): ",
            lambda x: float(x) if 0.0 <= float(x) <= 1.0 else float("err")
        )

    # ── Random Enchantment ────────────────────────────────────────────────────
    print("\nRandom Enchantment (RandomProperty / RandomSuffix):")
    print("  Items below level 20 use RandomProperty (prefix names like 'of the Bear').")
    print("  Items level 20+    use RandomSuffix   (suffix names, stat-point-scaled).")
    print("  [1] No random enchantments")
    print("  [2] Yes -- choose percentage of items to receive one")
    rand_mode = get_input("Choice: ", lambda x: int(x) if int(x) in [1, 2] else int("err"))
    rand_pct  = 0.0
    if rand_mode == 2:
        if not _ALL_ENCHANT_ENTRIES:
            print("  WARNING: No item_enchantment_template entries found. Disabling.")
            rand_mode = 1
        else:
            rand_pct = get_input(
                "Percentage of items to receive a random enchantment (0-100): ",
                lambda x: float(x) if 0.0 <= float(x) <= 100.0 else float("err")
            ) / 100.0

    # ── Item Count ───────────────────────────────────────────────────────────
    print("\nNote: Each item is ~3.6 KB in memory.")
    print("  Practical safe ceiling: ~400,000 items before RAM pressure.")
    print("  SQL file at 2,000 items ≈ 4 MB | at 50,000 items ≈ 100 MB")
    quantity = get_input(
        "How many items to generate? (recommended: 500–5000): ",
        lambda x: int(x) if int(x) > 0 else int("err")
    )

    # ── Anti-repetition Scheduler ────────────────────────────────────────────
    # Build a deck of (cat_idx, ilvl) pairs.
    # Strategy: for each unique category in the pool, assign ilvl slots spread
    # evenly across the requested range.  Shuffle the deck, pop items from it,
    # and reshuffle only when exhausted — this prevents clustering like
    # "ilvl 44 crossbow, ilvl 45 crossbow" while swords are untouched.

    ilvl_levels   = list(range(ilvl_range[0], ilvl_range[1] + 1))
    num_cats      = len(cat_indices)
    num_levels    = len(ilvl_levels)

    # One full "round" = every category × every ilvl (or a sampled subset if huge)
    # Cap a single round at 50k slots so the deck stays manageable.
    MAX_DECK_SLOTS = 50_000
    if num_cats * num_levels <= MAX_DECK_SLOTS:
        deck_pairs = [(ci, lvl) for ci in cat_indices for lvl in ilvl_levels]
    else:
        # Sample proportionally: pick enough ilvl slots per category
        slots_per_cat = max(1, MAX_DECK_SLOTS // num_cats)
        sampled_levels = ilvl_levels if len(ilvl_levels) <= slots_per_cat else \
            random.sample(ilvl_levels, slots_per_cat)
        deck_pairs = [(ci, lvl) for ci in cat_indices for lvl in sampled_levels]

    random.shuffle(deck_pairs)
    deck_index = 0  # rolling pointer — reshuffle when exhausted

    generated = 0
    skipped   = 0

    print(f"\n🚀 Generating {quantity:,} items across {num_cats} category types "
          f"and {num_levels} ilvl steps...\n")

    for _ in range(quantity):
        # Pull next (cat_idx, ilvl) from the anti-repetition deck
        if deck_index >= len(deck_pairs):
            random.shuffle(deck_pairs)
            deck_index = 0

        cat_idx, ilvl = deck_pairs[deck_index]
        deck_index += 1

        category    = CATEGORIES[cat_idx]
        quality_code = random.choice(quality_code_pool)

        lookup_key = (category["class"], category["subclass"], category["InventoryType"], quality_code)
        if lookup_key not in lookup_database:
            skipped += 1
            continue
        sheet = lookup_database[lookup_key]

        interpolated = get_interpolated_properties(sheet, ilvl, category["class"], category["subclass"], category["InventoryType"], quality_code)
        if not interpolated:
            skipped += 1
            continue

        # ── Math (identical to manual path) ──────────────────────────────────
        fuzz_factor  = random.uniform(1.0 - variance, 1.0 + variance)
        final_budget = int(interpolated["avg_budget"] * fuzz_factor)
        final_dps    = interpolated["avg_dps"] * fuzz_factor if category["class"] == 2 else 0.0

        dynamic_delay = 0
        if category["class"] == 2:
            base_delay    = subclass_delays.get(category["subclass"], category.get("delay", 2600))
            raw_delay     = random.uniform(base_delay * 0.85, base_delay * 1.15)
            dynamic_delay = int(round(raw_delay / 100) * 100)

        dmg_min, dmg_max = 0, 0
        if category["class"] == 2 and final_dps > 0:
            avg_damage  = final_dps * (dynamic_delay / 1000.0)
            spread_fuzz = random.uniform(-0.02, 0.02)
            dmg_min     = int(avg_damage * (0.70 + spread_fuzz))
            dmg_max     = int(avg_damage * (1.30 + spread_fuzz))
            final_dps   = round(((dmg_min + dmg_max) / 2) / (dynamic_delay / 1000.0), 2)

        if interpolated["stat_profiles"]:
            chosen_profile = random.choice(interpolated["stat_profiles"])
            db_stats_count = chosen_profile.get("num_stats", 2)
        else:
            db_stats_count = 2

        final_budget = max(1, int(final_budget * GLOBAL_BUDGET_NERF))

        if density_mode == 1:
            num_stats_to_roll = db_stats_count if db_stats_count > 0 else 2
        elif density_mode == 2:
            if ilvl < 30:   num_stats_to_roll = random.choices([1, 2], weights=[40, 60], k=1)[0]
            elif ilvl < 45: num_stats_to_roll = random.choices([2, 3], weights=[65, 35], k=1)[0]
            elif ilvl < 60: num_stats_to_roll = random.choices([2, 3], weights=[40, 60], k=1)[0]
            else:           num_stats_to_roll = random.choices([2, 3, 4], weights=[20, 65, 15], k=1)[0]
        elif density_mode == 3:
            num_stats_to_roll = chosen_density_count
        else:  # mode 4: random range
            num_stats_to_roll = random.randint(density_range_min, density_range_max)

        stats = {f"stat_type{i}": 0 for i in range(1, 7)}
        stats.update({f"stat_value{i}": 0 for i in range(1, 7)})

        if num_stats_to_roll > 0:
            # ── Blueprint auto-selection (no prompts) ─────────────────────
            item_blueprint_key = get_mass_blueprint(category)

            # Mail STR_TANK stops being relevant at level 40 when plate unlocks.
            # Reroll the item so the target count is always met.
            if (item_blueprint_key == "STR_TANK"
                    and category["class"] == 4
                    and category["subclass"] == 3
                    and ilvl > 40):
                continue

            blueprint = BLUEPRINTS.get(item_blueprint_key)

            if blueprint:
                pool, anchors, current_weights = (
                    list(blueprint["pool"]),
                    list(blueprint["anchors"]),
                    blueprint["weights"].copy()
                )
            else:
                pool, anchors, current_weights = [4, 7], [4, 7], {4: 100, 7: 100}

            if ilvl < 60:
                forbidden = {28, 30, 35, 36, 44}
                pool    = [s for s in pool    if s not in forbidden]
                anchors = [s for s in anchors if s not in forbidden]

            if (category["InventoryType"] in [17, 25, 26] or
                    category["subclass"] in [1, 5, 6, 8, 10]):
                pool    = [s for s in pool    if s not in [15, 48]]
                anchors = [s for s in anchors if s not in [15, 48]]

            if item_blueprint_key == "AGI_TANK":
                pool    = [s for s in pool    if s != 14]
                anchors = [s for s in anchors if s != 14]

            chosen_stats = []
            remaining_num = num_stats_to_roll
            for a in anchors:
                if a in pool and a not in chosen_stats:
                    chosen_stats.append(a)
                    remaining_num -= 1
                    if remaining_num <= 0:
                        break

            remaining_pool = [s for s in pool if s not in chosen_stats]
            if remaining_pool and remaining_num > 0:
                actual_extra = min(remaining_num, len(remaining_pool), 6 - len(chosen_stats))
                for _ in range(actual_extra):
                    valid_remaining = [s for s in remaining_pool if s not in chosen_stats]
                    if not valid_remaining:
                        break
                    w = [current_weights.get(s, 50) for s in valid_remaining]
                    chosen_stats.append(random.choices(valid_remaining, weights=w, k=1)[0])

            num_active_stats = len(chosen_stats)
            shares = [1.0 / num_active_stats for _ in range(num_active_stats)]
            if dist_mode == 2 and num_active_stats > 1:
                shares = [max(0.01, s + random.uniform(-skew_factor / 100.0, skew_factor / 100.0))
                          for s in shares]
                s_sum  = sum(shares)
                shares = [s / s_sum for s in shares]

            allocated_values = [max(1, int(final_budget * s)) for s in shares]
            remainder        = final_budget - sum(allocated_values)
            if remainder != 0 and allocated_values:
                max_idx = allocated_values.index(max(allocated_values))
                allocated_values[max_idx] += remainder

            for idx, stat_type in enumerate(chosen_stats):
                stats[f"stat_type{idx+1}"]  = stat_type
                stats[f"stat_value{idx+1}"] = allocated_values[idx]

        # ── Cosmetics ────────────────────────────────────────────────────────
        m_key = (category["class"], category["subclass"], quality_code)
        if m_key in material_library:
            chosen_pair  = random.choice(material_library[m_key])
            item_material = chosen_pair['Material']
            item_sheath   = get_sheathe_type(category)
        else:
            fallback_key   = (category["class"], category["subclass"])
            possible_keys  = [k for k in material_library if k[0:2] == fallback_key]
            if possible_keys:
                chosen_pair   = random.choice(random.choice([material_library[k] for k in possible_keys]))
                item_material = chosen_pair['Material']
                item_sheath   = get_sheathe_type(category)
            else:
                item_material, item_sheath = 1, 0

        cat_keys       = (category["class"], category["subclass"])
        generated_name = generate_item_name((cat_keys[0], cat_keys[1], lookup_key[2]))
        display_obj    = get_appropriate_display_id(cat_keys, ilvl, quality_code, category["InventoryType"])
        req_level      = get_appropriate_req_level(cursor, ilvl, quality_code)

        if category['InventoryType'] in (2, 11, 23):  # Neck, Ring, Off-hand frill
            avg_armor  = 0.0
        else:
            avg_armor  = get_scaled_armor(lookup_database, category, quality_code, ilvl, sheet) if category['class'] == 4 else 0.0
        fuzz_factor2   = random.uniform(1.0 - variance, 1.0 + variance)
        final_armor    = int(avg_armor * fuzz_factor2) if avg_armor > 0 else 0
        final_block    = max(1, int(get_scaled_block(lookup_database, category, quality_code, ilvl, sheet) * fuzz_factor2)) if category['InventoryType'] == 14 else 0
        if bonding_mode == 1:   item_bonding = 2
        elif bonding_mode == 2: item_bonding = 1
        else:                   item_bonding = 2 if random.random() < boe_ratio else 1
        item_rand_prop = 0
        item_rand_suf  = 0
        if rand_mode == 2 and _ALL_ENCHANT_ENTRIES and random.random() < rand_pct:
            ench_entry = random.choice(_ALL_ENCHANT_ENTRIES)
            if ilvl < 20:
                item_rand_prop = ench_entry
            else:
                item_rand_suf  = ench_entry

        base_sell_price  = calculate_item_sell_price(lookup_database, category, quality_code, ilvl)
        final_sell_price = int(base_sell_price * fuzz_factor2)

        internal_memory.append({
            "config": category, "quality": quality_code, "ilvl": ilvl,
            "name": generated_name, "displayid": display_obj.get("id"),
            "display_source": display_obj,
            "dmg_min": dmg_min, "dmg_max": dmg_max,
            "delay": dynamic_delay, "dps": final_dps,
            "armor": final_armor, "block": final_block,
            "stats": stats, "budget": final_budget,
            "required_level": req_level, "Material": item_material,
            "sheath": item_sheath, "sell_price": final_sell_price,
            "bonding": item_bonding,
            "rand_prop": item_rand_prop, "rand_suf": item_rand_suf,
            "itemset": 0,
            "spellid": 0, "spelltrigger": 0, "spellcharges": 0, "spellcooldown": -1
        })
        generated += 1

        # Lightweight progress ticker every 100 items
        if generated % 100 == 0 or generated == quantity:
            pct = generated / quantity * 100
            bar = "█" * (generated * 20 // quantity) + "░" * (20 - generated * 20 // quantity)
            print(f"\r  [{bar}] {generated:>5}/{quantity}  ({pct:.0f}%)  — {skipped} skipped", end="", flush=True)

    print(f"\n\n✅ Mass Creation complete!  Generated: {generated:,}  |  Skipped: {skipped}")
    print(f"   Total items in memory: {len(internal_memory):,}")


# ═════════════════════════════════════════════════════════════════════════════


# =============================================================================
# LOOT ASSIGNMENT ENGINE
# Reads existing creature_loot_template and reference_loot_template, then
# assigns every item in internal_memory to level-appropriate mobs.
#
# Open-world mobs   -> direct rows in creature_loot_template (independent roll)
# Dungeon/raid boss -> rows inserted into the boss's existing reference group
#                      in reference_loot_template so the new item competes with
#                      existing drops in the same roll pool rather than getting
#                      a free separate roll.
# =============================================================================

def run_loot_assignment():
    if not internal_memory:
        print("\n  No items in memory. Generate some items first.")
        return

    print("\n" + "="*57)
    print("  LOOT ASSIGNMENT ENGINE")
    print("="*57)

    # -- Configuration ---------------------------------------------------------
    print("\nDrop chances (press Enter to keep default):")

    def _pct(prompt, default):
        raw = input(f"  {prompt} [{default}%]: ").strip()
        if raw == "":
            return default
        try:
            v = float(raw)
            if 0 < v <= 100:
                return v
        except ValueError:
            pass
        print(f"  Invalid value, using {default}%")
        return default

    chance_green  = _pct("Green  (Uncommon)", 10.0)
    chance_blue   = _pct("Blue   (Rare)",      5.0)
    chance_purple = _pct("Purple (Epic)",       1.0)
    CHANCE_MAP    = {2: chance_green, 3: chance_blue, 4: chance_purple}

    print("\nMob scope:")
    print("  [1] Open-world mobs only")
    print("  [2] Dungeon / raid bosses only")
    print("  [3] Both")
    scope = get_input("Choice: ",
        lambda x: int(x) if x in ["1", "2", "3"] else int("err"))
    include_openworld = scope in (1, 3)
    include_instances = scope in (2, 3)

    raw_tol = input("  iLvl tolerance for boss matching [10]: ").strip()
    try:
        ilvl_tol = int(raw_tol) if raw_tol else 10
        if ilvl_tol <= 0:
            ilvl_tol = 10
    except ValueError:
        ilvl_tol = 10

    # -- Re-open a fresh DB connection ----------------------------------------
    # The main cursor may have been closed after generation.
    print("\n -> Connecting to database...")
    try:
        loot_conn   = mysql.connector.connect(**DB_CONFIG)
        loot_cursor = loot_conn.cursor(dictionary=True)
    except Exception as e:
        print(f"  ERROR: Could not connect to database: {e}")
        return

    # -- Classify maps: open-world vs instanced --------------------------------
    # Maps 0 (Eastern Kingdoms) and 1 (Kalimdor) are the open world.
    # Everything else that appears in the creature spawn table is an instance.
    # We try map_dbc first for authoritative map_type data:
    #   0 = World, 1 = Instance, 2 = Raid, 3 = Battleground, 4 = Arena
    print(" -> Classifying maps...")
    try:
        loot_cursor.execute("SELECT DISTINCT map FROM creature")
        all_maps = {row["map"] for row in loot_cursor.fetchall()}

        try:
            loot_cursor.execute("SELECT entry, map_type FROM map_dbc")
            map_types     = {r["entry"]: r["map_type"] for r in loot_cursor.fetchall()}
            instance_maps = {m for m in all_maps if map_types.get(m, 0) in (1, 2)}
        except Exception:
            # Fallback: treat everything except 0 and 1 as instanced
            instance_maps = all_maps - {0, 1}

        openworld_maps = all_maps - instance_maps
    except Exception as e:
        print(f"  WARNING: Map classification failed ({e}). Defaulting to ID heuristic.")
        instance_maps  = set()
        openworld_maps = {0, 1}

    # -- Load creature templates -----------------------------------------------
    # One row per template; maps aggregated with GROUP_CONCAT.
    print(" -> Loading creature templates...")
    loot_cursor.execute("""
        SELECT
            ct.entry,
            ct.name,
            ct.minlevel,
            ct.maxlevel,
            ct.lootid,
            ct.rank,
            GROUP_CONCAT(DISTINCT c.map) AS maps
        FROM creature_template ct
        JOIN creature c ON c.id1 = ct.entry
        WHERE ct.lootid != 0
          AND ct.minlevel > 0
        GROUP BY ct.entry
    """)
    all_creatures = loot_cursor.fetchall()

    openworld_creatures = []
    boss_creatures      = []

    for row in all_creatures:
        maps        = set(int(m) for m in (row["maps"] or "").split(",") if m.strip())
        in_instance = bool(maps & instance_maps)
        in_openworld = bool(maps & openworld_maps)
        is_boss     = int(row["rank"]) in (3, 4)   # Boss / Rare-Elite

        if include_instances and in_instance and is_boss:
            boss_creatures.append(row)
        if include_openworld and in_openworld and not is_boss:
            openworld_creatures.append(row)

    print(f"    Open-world mobs found   : {len(openworld_creatures):,}")
    print(f"    Dungeon/raid bosses found: {len(boss_creatures):,}")

    # -- Analyse boss reference loot tables ------------------------------------
    # For each boss we find rows in creature_loot_template where Reference != 0.
    # Those Reference values point to reference_loot_template entries.
    # We compute the average ilvl of real items in each reference group so we
    # can match generated items to the closest-ilvl group.
    #
    # Critically: we filter out rlt.Reference != 0 rows because AzerothCore has
    # a known bug where reference-of-reference entries inside a group are silently
    # skipped by the loot engine -- inserting there would have no effect.
    print(" -> Analysing boss reference loot tables...")

    boss_lootids = {int(row["lootid"]) for row in boss_creatures}
    lootid_to_refs = {}   # lootid -> [{ref_entry, group_id, avg_ilvl}]

    if boss_lootids:
        ids_str = ",".join(str(x) for x in boss_lootids)
        loot_cursor.execute(f"""
            SELECT
                clt.Entry      AS loot_id,
                clt.Reference  AS ref_entry,
                clt.GroupId    AS group_id,
                AVG(it.ItemLevel) AS avg_ilvl,
                COUNT(rlt.Item)   AS item_count
            FROM creature_loot_template clt
            JOIN reference_loot_template rlt ON rlt.Entry = clt.Reference
            JOIN item_template it ON it.entry = rlt.Item
            WHERE clt.Entry IN ({ids_str})
              AND clt.Reference != 0
              AND rlt.Reference = 0
              AND rlt.Item != 0
            GROUP BY clt.Entry, clt.Reference, clt.GroupId
        """)
        for row in loot_cursor.fetchall():
            lid = int(row["loot_id"])
            if lid not in lootid_to_refs:
                lootid_to_refs[lid] = []
            lootid_to_refs[lid].append({
                "ref_entry": int(row["ref_entry"]),
                "group_id":  int(row["group_id"]),
                "avg_ilvl":  float(row["avg_ilvl"] or 0),
            })

    # -- Collect items already assigned to avoid duplicates --------------------
    print(" -> Checking for existing loot assignments...")
    assigned_item_ids = set()
    try:
        loot_cursor.execute(
            "SELECT DISTINCT Item FROM creature_loot_template WHERE Item > 0")
        assigned_item_ids.update(r["Item"] for r in loot_cursor.fetchall())
        loot_cursor.execute(
            "SELECT DISTINCT Item FROM reference_loot_template WHERE Item > 0")
        assigned_item_ids.update(r["Item"] for r in loot_cursor.fetchall())
    except Exception as e:
        print(f"  WARNING: Could not load existing assignments ({e}).")

    # The item entry IDs we generated start at start_entry_id (same as SQL output)
    start_id = get_next_entry_id(cursor)

    # -- Index open-world creatures by level for fast lookup ------------------
    ow_by_level = {}   # level -> [lootid, ...]  (deduped per level)
    for row in openworld_creatures:
        for lvl in range(int(row["minlevel"]), int(row["maxlevel"]) + 1):
            ow_by_level.setdefault(lvl, set()).add(int(row["lootid"]))

    # -- Build assignment plan -------------------------------------------------
    print(" -> Planning assignments...")

    openworld_inserts = []   # (loot_entry, item_id, chance, comment)
    reference_inserts = []   # (ref_entry, item_id, group_id, chance, comment)
    unassigned        = []   # (item_id, name, ilvl)

    for idx, item in enumerate(internal_memory):
        item_id      = start_id + idx
        item_ilvl    = int(item["ilvl"])
        item_quality = int(item["quality"])
        item_name    = item["name"]
        item_chance  = CHANCE_MAP.get(item_quality, 5.0)

        if item_id in assigned_item_ids:
            continue

        assigned = False

        # Open-world: find mobs whose level range overlaps [req_lvl, req_lvl+10]
        if include_openworld:
            req_lvl  = int(item.get("required_level", max(1, item_ilvl - 5)))
            hit_lootids = set()
            for lvl in range(req_lvl, req_lvl + 11):
                hit_lootids.update(ow_by_level.get(lvl, set()))
            for loot_entry in hit_lootids:
                openworld_inserts.append((
                    loot_entry, item_id, item_chance,
                    f"{item_name} iLvl{item_ilvl} (generated)"
                ))
            if hit_lootids:
                assigned = True

        # Boss: find the reference group whose avg_ilvl is closest to item_ilvl
        # within tolerance. Prefer the single best match per item to avoid
        # flooding every boss with every item -- one boss per generated item.
        if include_instances:
            best_ref   = None
            best_delta = float("inf")
            for boss_row in boss_creatures:
                lootid = int(boss_row["lootid"])
                for ref in lootid_to_refs.get(lootid, []):
                    delta = abs(ref["avg_ilvl"] - item_ilvl)
                    if delta <= ilvl_tol and delta < best_delta:
                        best_delta = delta
                        best_ref   = ref

            if best_ref:
                reference_inserts.append((
                    best_ref["ref_entry"],
                    item_id,
                    best_ref["group_id"],
                    item_chance,
                    f"{item_name} iLvl{item_ilvl} (generated)"
                ))
                assigned = True

        if not assigned:
            unassigned.append((item_id, item_name, item_ilvl))

    # -- Write loot_assignments.sql --------------------------------------------
    loot_filename = "loot_assignments.sql"
    print(f"\n -> Writing {loot_filename}...")

    with open(loot_filename, "w", encoding="utf-8") as f:
        f.write("-- ===========================================================\n")
        f.write("-- AzerothCore Loot Assignments -- generated by Item Generator\n")
        f.write(f"-- Open-world rows : {len(openworld_inserts):,}\n")
        f.write(f"-- Reference rows  : {len(reference_inserts):,}\n")
        f.write(f"-- Unassigned items: {len(unassigned):,}\n")
        f.write("-- ===========================================================\n\n")

        if openworld_inserts:
            f.write("-- Open-world mob loot (direct creature_loot_template rows)\n")
            f.write("-- GroupId=0: each item rolls independently.\n\n")
            for (loot_entry, item_id, chance, comment) in openworld_inserts:
                f.write(
                    f"INSERT IGNORE INTO `creature_loot_template` "
                    f"(`Entry`,`Item`,`Reference`,`Chance`,`QuestRequired`,"
                    f"`LootMode`,`GroupId`,`MinCount`,`MaxCount`,`Comment`) "
                    f"VALUES ({loot_entry},{item_id},0,{chance},0,1,0,1,1,"
                    f"'{comment}');\n"
                )
            f.write("\n")

        if reference_inserts:
            f.write("-- Dungeon/raid boss loot (reference_loot_template rows)\n")
            f.write("-- Inserted into boss's existing reference group so the item\n")
            f.write("-- competes with existing drops in the same roll pool.\n\n")
            for (ref_entry, item_id, group_id, chance, comment) in reference_inserts:
                f.write(
                    f"INSERT IGNORE INTO `reference_loot_template` "
                    f"(`Entry`,`Item`,`Reference`,`Chance`,`QuestRequired`,"
                    f"`LootMode`,`GroupId`,`MinCount`,`MaxCount`,`Comment`) "
                    f"VALUES ({ref_entry},{item_id},0,{chance},0,1,{group_id},1,1,"
                    f"'{comment}');\n"
                )
            f.write("\n")

        if unassigned:
            f.write("-- WARNING: items below could not be matched to any mob.\n")
            f.write("-- Possible causes: ilvl outside range of available mobs,\n")
            f.write("-- no boss reference groups within tolerance, or item already\n")
            f.write("-- assigned from a previous run.\n")
            for (item_id, name, ilvl) in unassigned:
                f.write(f"--   entry {item_id}: {name} (iLvl {ilvl})\n")

    loot_cursor.close()
    loot_conn.close()

    # -- Summary ---------------------------------------------------------------
    print(f"\n  Open-world assignments  : {len(openworld_inserts):>6,}")
    print(f"  Reference  assignments  : {len(reference_inserts):>6,}")
    if unassigned:
        print(f"  Unassigned items        : {len(unassigned):>6,}  "
              f"(listed as comments in {loot_filename})")
    print(f"\n  Saved -> {loot_filename}")


def run_itemset_creation():
    """Groups items already sitting in internal_memory into an ItemSet.
    Tags each member item with item_template.itemset = set_id and records
    the set's spell bonuses for later ItemSet.dbc CSV export."""
    if not internal_memory:
        print("\n  No items in memory. Generate some items first.")
        return

    print("\n" + "="*57)
    print("  ITEM SET CREATOR")
    print("="*57)

    print(f"\n  {len(internal_memory)} item(s) currently in memory:\n")
    for i, item in enumerate(internal_memory, start=1):
        c = item["config"]
        tag = f"  [already in Set {item['itemset']}]" if item.get("itemset") else ""
        print(f"   [{i:>4}] {item['name']}  (iLvl {item['ilvl']}, Q{item['quality']}, {c['name']}){tag}")

    print("\nEnter the numbers of the items belonging to this set (2-17 items),")
    print("comma-separated, e.g. '3,7,12':")
    while True:
        raw = input("  Items: ").strip()
        try:
            picks = sorted(set(int(x.strip()) for x in raw.split(",") if x.strip()))
        except ValueError:
            print("  Invalid input. Use comma-separated numbers.")
            continue
        if not (2 <= len(picks) <= 17):
            print("  A set needs between 2 and 17 items.")
            continue
        if any(p < 1 or p > len(internal_memory) for p in picks):
            print("  One or more numbers are out of range.")
            continue
        break
    item_indices = [p - 1 for p in picks]  # zero-based positions into internal_memory

    set_id = get_input(
        "\nEnter Item Set ID (must be unique -- matches ItemSet.dbc ID and "
        "item_template.itemset): ",
        lambda x: int(x) if int(x) > 0 else int("err")
    )
    if any(s["id"] == set_id for s in generated_itemsets):
        print(f"  WARNING: Set ID {set_id} was already used earlier in this session. Aborting.")
        return

    set_name = input("Enter Set Name (e.g. 'Sorvaxis Battlegear'): ").strip()
    if not set_name:
        set_name = f"Unnamed Set {set_id}"

    print(f"\nDefine set bonuses (spell triggered once N of the {len(picks)} items are worn).")
    print("Up to 8 bonus tiers. Leave Spell ID empty to stop adding bonuses.")
    spells = []
    for slot in range(1, 9):
        raw_spell = input(f"  Bonus {slot} -- Spell ID (Enter to stop): ").strip()
        if raw_spell == "":
            break
        try:
            spell_id = int(raw_spell)
        except ValueError:
            print("  Invalid Spell ID, skipping this slot.")
            continue
        threshold = get_input(
            f"  Bonus {slot} -- Threshold (2-{len(picks)} items required): ",
            lambda x: int(x) if 2 <= int(x) <= len(picks) else int("err")
        )
        spells.append((threshold, spell_id))

    # Tag every selected item with this set's ID (written to item_template.itemset on export)
    for i in item_indices:
        internal_memory[i]["itemset"] = set_id

    generated_itemsets.append({
        "id": set_id, "name": set_name,
        "item_indices": item_indices, "spells": spells
    })

    print(f"\n✅ Set '{set_name}' (ID {set_id}) created with {len(item_indices)} items "
          f"and {len(spells)} bonus tier(s).")


while True:
    print("\n--- NEW ITEM GENERATION BATCH ---")

    # ── TIER 1: Top-level group ──────────────────────────────────────────────
    print("Select Category Group:")
    print("  [1] Weapons")
    print("  [2] Armor")
    print("  [3] All of the above (random per item)")
    print("  [4] ⚡ Mass Creation (auto-populate, no archetype prompts)")
    print("  [5] 🎯 Assign Loot to Mobs (uses items already in memory)")
    print("  [6] 🎁 Create Item Set (group items already in memory)")
    group_choice = get_input("Choice: ", lambda x: int(x) if x in ['1', '2', '3', '4', '5', '6'] else int("err"))

    if group_choice == 4:
        run_mass_creation()
        if get_input("\nAdd another batch? (y/n): ", lambda x: x.lower() if x.lower() in ['y', 'n'] else int("err")) == 'n':
            break
        continue

    if group_choice == 5:
        run_loot_assignment()
        if get_input("\nReturn to menu? (y/n): ", lambda x: x.lower() if x.lower() in ['y', 'n'] else int("err")) == 'n':
            break
        continue

    if group_choice == 6:
        run_itemset_creation()
        if get_input("\nReturn to menu? (y/n): ", lambda x: x.lower() if x.lower() in ['y', 'n'] else int("err")) == 'n':
            break
        continue

    # cat_pool = list of (cat_idx, blueprint_key, blueprint_candidates) resolved per item
    # We first collect a pool of candidate cat_indices, then resolve blueprints.

    cat_indices = []  # will hold one or more CATEGORIES keys

    if group_choice == 3:  # ALL: every weapon + every armor slot
        cat_indices = list(CATEGORIES.keys())

    elif group_choice == 1:  # WEAPONS
        # ── TIER 2: Weapon group ──
        print("\nSelect Weapon Group:")
        print("  [1] 1H Weapons")
        print("  [2] 2H Weapons")
        print("  [3] Ranged")
        print("  [4] All of the above (random per item)")
        wpn_group = get_input("Select Group: ", lambda x: int(x) if x in ['1','2','3','4'] else int("err"))

        if wpn_group == 4:  # ALL weapon groups
            cat_indices = [4, 6, 7, 15, 2,   # 1H
                           5, 9, 10, 8, 3,    # 2H
                           12, 11, 13, 14, 1] # Ranged

        elif wpn_group == 1:  # 1H
            cat_indices = pick_cat_indices(
                "Select 1H Weapon:",
                {1: 4, 2: 6, 3: 7, 4: 15, 5: 2})

        elif wpn_group == 2:  # 2H
            cat_indices = pick_cat_indices(
                "Select 2H Weapon:",
                {1: 5, 2: 9, 3: 10, 4: 8, 5: 3})

        elif wpn_group == 3:  # Ranged
            cat_indices = pick_cat_indices(
                "Select Ranged Weapon:",
                {1: 12, 2: 11, 3: 13, 4: 14, 5: 1})

    else:  # ARMOR (group_choice == 2)
        # ── TIER 2: Armor material ──
        print("\nSelect Armor Material:")
        print("  [1] Cloth")
        print("  [2] Leather")
        print("  [3] Mail")
        print("  [4] Plate")
        print("  [5] Miscellaneous")
        print("  [6] All of the above (random per item)")
        mat = get_input("Select Material: ", lambda x: int(x) if x in ['1','2','3','4','5','6'] else int("err"))

        if mat == 6:  # ALL armor
            cat_indices = list(range(16, 53))  # Cloth(16-23), Leather(24-31), Mail(32-39), Plate(40-47), Misc(48-52)

        elif mat <= 4:
            base = {1: 16, 2: 24, 3: 32, 4: 40}[mat]
            mat_name = {1: "Cloth", 2: "Leather", 3: "Mail", 4: "Plate"}[mat]
            # ── TIER 3: Armor slot ──
            cat_indices = pick_cat_indices(
                f"Select {mat_name} Slot:",
                {1: base+0, 2: base+1, 3: base+2, 4: base+3,
                 5: base+4, 6: base+5, 7: base+6, 8: base+7})

        elif mat == 5:  # Miscellaneous
            cat_indices = pick_cat_indices(
                "Select Miscellaneous Slot:",
                {1: 48, 2: 49, 3: 50, 4: 51, 5: 52})

    # ── Blueprint resolution ─────────────────────────────────────────────────
    # If exactly one category was chosen → ask blueprint once (original behaviour).
    # If multiple categories were chosen → ask blueprint once per unique category type,
    # then map each cat_idx to its resolved (key, candidates) pair.

    unique_cat_indices = list(dict.fromkeys(cat_indices))  # deduplicate, preserve order

    if len(unique_cat_indices) == 1:
        # Single category: ask once, apply to all items
        sole_category = CATEGORIES[unique_cat_indices[0]]
        chosen_blueprint_key, chosen_blueprint_candidates = resolve_blueprint_for_category(sole_category)
        # Build a flat lookup: cat_idx -> (key, candidates)
        blueprint_map = {unique_cat_indices[0]: (chosen_blueprint_key, chosen_blueprint_candidates)}
    else:
        # Multiple categories: ask blueprint for each unique category type
        blueprint_map = {}
        seen_signatures = {}  # (cls, subc, invtype) -> resolved result, avoid asking twice for same type
        for ci in unique_cat_indices:
            cat = CATEGORIES[ci]
            sig = (cat["class"], cat["subclass"], cat["InventoryType"])
            if sig not in seen_signatures:
                print(f"\n  ── Blueprint for: {cat['name']} ──")
                result = resolve_blueprint_for_category(cat)
                seen_signatures[sig] = result
            blueprint_map[ci] = seen_signatures[sig]

    # ── Shared generation parameters ────────────────────────────────────────
    print("\nAvailable Qualities:")
    for k, v in QUALITIES.items(): print(f"  [{k}] {v['name']}")
    selected_quality_entry = QUALITIES[get_input("Select Item Quality (Number): ", lambda x: int(x) if int(x) in QUALITIES else int("err"))]
    quality_code_pool = selected_quality_entry["code"] if selected_quality_entry["multi"] else [selected_quality_entry["code"]]

    print("\nEnter Target Item Level (e.g., '85' or '50-85'):")
    ilvl_range = get_input("Choice: ", parse_ilvl_input)
    variance = get_input("Enter Budget Quality Variance % (0 to 25): ", lambda x: float(x) / 100.0 if 0 <= float(x) <= 25 else float("err"))

    print("\nSelect Stat Distribution Allocation Profile:")
    print("  [1] Even Split\n  [2] Randomly Varied Split")
    dist_mode = get_input("Select Profile Mode (1 or 2): ", lambda x: int(x) if int(x) in [1, 2] else int("err"))
    skew_factor = get_input("Enter Max Deviation %: ", lambda x: float(x) if 0 <= float(x) <= 100 else float("err")) if dist_mode == 2 else 0.0

    print("\nSelect Stat Slot Density Rule:")
    print("  [1] Database-Driven\n  [2] Progressive Blizzlike\n  [3] Explicit Manual Count\n  [4] Random Range")
    density_mode = get_input("Select Density Rule (1, 2, 3, or 4): ", lambda x: int(x) if int(x) in [1, 2, 3, 4] else int("err"))
    chosen_density_count = get_input("Enter exact count (1 to 6): ", lambda x: int(x) if 1 <= int(x) <= 6 else int("err")) if density_mode == 3 else 0
    density_range_min, density_range_max = 0, 0
    if density_mode == 4:
        density_range_min = get_input(
            "Enter minimum stat count (1-6): ",
            lambda x: int(x) if 1 <= int(x) <= 6 else int("err")
        )
        density_range_max = get_input(
            f"Enter maximum stat count ({density_range_min}-6): ",
            lambda x: int(x) if density_range_min <= int(x) <= 6 else int("err")
        )

    print("\nItem Binding:")
    print("  [1] Bind on Equip (BoE)")
    print("  [2] Bind on Pickup (BoP)")
    print("  [3] Both (mixed)")
    bonding_mode = get_input("Choice: ", lambda x: int(x) if int(x) in [1, 2, 3] else int("err"))
    boe_ratio = 0.5
    if bonding_mode == 3:
        boe_ratio = get_input(
            "Proportion of BoE items (0.0 = all BoP, 1.0 = all BoE): ",
            lambda x: float(x) if 0.0 <= float(x) <= 1.0 else float("err")
        )

    print("\nRandom Enchantment (RandomProperty / RandomSuffix):")
    print("  Items below level 20 use RandomProperty (prefix names like 'of the Bear').")
    print("  Items level 20+    use RandomSuffix   (suffix names, stat-point-scaled).")
    print("  [1] No random enchantments")
    print("  [2] Yes -- choose percentage of items to receive one")
    rand_mode = get_input("Choice: ", lambda x: int(x) if int(x) in [1, 2] else int("err"))
    rand_pct  = 0.0
    if rand_mode == 2:
        if not _ALL_ENCHANT_ENTRIES:
            print("  WARNING: No item_enchantment_template entries found. Disabling.")
            rand_mode = 1
        else:
            rand_pct = get_input(
                "Percentage of items to receive a random enchantment (0-100): ",
                lambda x: float(x) if 0.0 <= float(x) <= 100.0 else float("err")
            ) / 100.0

    print("\nItem Spell (On-Equip Aura / On-Use Effect):")
    print("  [1] No spell")
    print("  [2] Yes -- On Equip (passive aura while worn, e.g. proc/stat effect)")
    print("  [3] Yes -- On Use (activated ability with charges/cooldown)")
    spell_mode = get_input("Choice: ", lambda x: int(x) if int(x) in [1, 2, 3] else int("err"))
    batch_spell_id = 0
    batch_spell_trigger = 0
    batch_spell_charges = 0
    batch_spell_cooldown = -1
    spell_pct = 100.0
    if spell_mode in (2, 3):
        batch_spell_id = get_input("Enter Spell ID: ", lambda x: int(x) if int(x) > 0 else int("err"))
        batch_spell_trigger = 1 if spell_mode == 2 else 0  # 1=On Equip, 0=On Use
        if spell_mode == 3:
            batch_spell_charges = get_input(
                "Charges (0 = unlimited uses, >0 = item is consumed after N uses): ",
                lambda x: int(x) if int(x) >= 0 else int("err")
            )
            batch_spell_cooldown = get_input(
                "Cooldown in ms (-1 = use the spell's own cooldown from spell.dbc): ",
                lambda x: int(x) if int(x) >= -1 else int("err")
            )
        spell_pct = get_input(
            "Percentage of items in this batch that receive the spell (0-100): ",
            lambda x: float(x) if 0.0 <= float(x) <= 100.0 else float("err")
        )

    quantity = get_input("\nHow many items?: ", lambda x: int(x) if int(x) > 0 else int("err"))

    available_levels = list(range(ilvl_range[0], ilvl_range[1] + 1))
    random.shuffle(available_levels)

    for _ in range(quantity):
        # ── Per-item: draw quality ──
        quality_code = random.choice(quality_code_pool)

        # ── Per-item: draw iLevel (deck shuffles on exhaust) ──
        if not available_levels:
            available_levels = list(range(ilvl_range[0], ilvl_range[1] + 1))
            random.shuffle(available_levels)
        ilvl = available_levels.pop()

        # ── Per-item: pick a category from the pool ──
        cat_idx = random.choice(cat_indices)
        category = CATEGORIES[cat_idx]

        # ── Per-item: resolve blueprint ──
        item_blueprint_key_pair = blueprint_map.get(cat_idx)
        if item_blueprint_key_pair:
            chosen_blueprint_key, chosen_blueprint_candidates = item_blueprint_key_pair
        else:
            chosen_blueprint_key, chosen_blueprint_candidates = "STR_DPS", ["STR_DPS"]

        if chosen_blueprint_key == "__ALL__":
            item_blueprint_key = random.choice(chosen_blueprint_candidates)
        else:
            item_blueprint_key = chosen_blueprint_key

        lookup_key = (category["class"], category["subclass"], category["InventoryType"], quality_code)
        if lookup_key not in lookup_database:
            print(f"⚠️ Configuration mapping missing for {category['name']} (Q{quality_code}). Skipping item.")
            continue
        sheet = lookup_database[lookup_key]

        # 6. GET INTERPOLATION for THIS specific iLevel
        interpolated = get_interpolated_properties(sheet, ilvl, category["class"], category["subclass"], category["InventoryType"], quality_code)
        
        if not interpolated:
            print(f"⚠️ Interpolation frame fault at iLvl {ilvl}. Skipping.")
            continue

        # --- ALL MATH LOGIC IS NOW INSIDE THIS ONE LOOP ---
        fuzz_factor = random.uniform(1.0 - variance, 1.0 + variance)
        final_budget = int(interpolated["avg_budget"] * fuzz_factor)
        final_dps = interpolated["avg_dps"] * fuzz_factor if category["class"] == 2 else 0.0
        
        dynamic_delay = 0
        if category["class"] == 2:
            # Fallback to the hardcoded CATEGORIES configuration value if database lacks tracking 
            base_delay = subclass_delays.get(category["subclass"], category.get("delay", 2600))
            raw_delay = random.uniform(base_delay * 0.85, base_delay * 1.15)
            dynamic_delay = int(round(raw_delay / 100) * 100) # Snap to clean hundredths

        # --- DYNAMIC SWING RANGE DAMAGE CALCULATIONS ---
        dmg_min, dmg_max = 0, 0
        if category["class"] == 2 and final_dps > 0:
            # CHANGED: Replaced static category["delay"] with our randomized dynamic_delay
            avg_damage = (final_dps * (dynamic_delay / 1000.0))
            spread_fuzz = random.uniform(-0.02, 0.02)
            dmg_min = int(avg_damage * (0.70 + spread_fuzz))
            dmg_max = int(avg_damage * (1.30 + spread_fuzz))
            final_dps = round(((dmg_min + dmg_max) / 2) / (dynamic_delay / 1000.0), 2)

        if interpolated["stat_profiles"]:
            chosen_profile = random.choice(interpolated["stat_profiles"])
            db_stats_count = chosen_profile.get("num_stats", 2)
        else:
            db_stats_count = 2

        final_budget = max(1, int(final_budget * GLOBAL_BUDGET_NERF))

        if density_mode == 1:
            num_stats_to_roll = db_stats_count if db_stats_count > 0 else 2
        elif density_mode == 2:
            if ilvl < 30: num_stats_to_roll = random.choices([1, 2], weights=[40, 60], k=1)[0]
            elif ilvl < 45: num_stats_to_roll = random.choices([2, 3], weights=[65, 35], k=1)[0]
            elif ilvl < 60: num_stats_to_roll = random.choices([2, 3], weights=[40, 60], k=1)[0]
            else: num_stats_to_roll = random.choices([2, 3, 4], weights=[20, 65, 15], k=1)[0]
        elif density_mode == 3:
            num_stats_to_roll = chosen_density_count
        else:  # mode 4: random range
            num_stats_to_roll = random.randint(density_range_min, density_range_max)

        stats = {f"stat_type{i}": 0 for i in range(1, 7)}
        stats.update({f"stat_value{i}": 0 for i in range(1, 7)})

        if num_stats_to_roll > 0:
            if item_blueprint_key and item_blueprint_key in BLUEPRINTS:
                blueprint = BLUEPRINTS[item_blueprint_key]
                pool, anchors, current_weights = list(blueprint["pool"]), list(blueprint["anchors"]), blueprint["weights"].copy()
            else:
                profile_key = (category["class"], category["subclass"], category["InventoryType"])
                profile = archetype_profiles.get(profile_key)
                if profile and profile["weights"]:
                    archs, weights = list(profile["weights"].keys()), list(profile["weights"].values())
                    chosen_arch = random.choices(archs, weights=weights, k=1)[0]
                    pool, anchors, current_weights = list(profile["stats"].get(chosen_arch, [4, 7])), [4, 7], {s: 100 for s in [4, 7]}
                else:
                    pool, anchors, current_weights = [4, 7], [4, 7], {4: 100, 7: 100}

            if ilvl < 60:
                forbidden = {28, 30, 35, 36, 44}
                pool = [s for s in pool if s not in forbidden]
                anchors = [s for s in anchors if s not in forbidden]

            if category["InventoryType"] in [17, 25, 26] or category["subclass"] in [1, 5, 6, 8, 10]:
                pool = [s for s in pool if s not in [15, 48]]
                anchors = [s for s in anchors if s not in [15, 48]]
            if item_blueprint_key == "AGI_TANK":
                pool = [s for s in pool if s != 14]
                anchors = [s for s in anchors if s != 14]

            chosen_stats = []
            if anchors and num_stats_to_roll > 0:
                for a in anchors:
                    if a in pool and a not in chosen_stats:
                        chosen_stats.append(a)
                        num_stats_to_roll -= 1
                        if num_stats_to_roll <= 0: break
            
            remaining_pool = [s for s in pool if s not in chosen_stats]
            if remaining_pool and num_stats_to_roll > 0:
                actual_extra = min(num_stats_to_roll, len(remaining_pool), 6 - len(chosen_stats))
                for _ in range(actual_extra):
                    valid_remaining = [s for s in remaining_pool if s not in chosen_stats]
                    if not valid_remaining: break
                    weights = [current_weights.get(s, 50) for s in valid_remaining]
                    chosen_stat = random.choices(valid_remaining, weights=weights, k=1)[0]
                    chosen_stats.append(chosen_stat)
                
            num_active_stats = len(chosen_stats)
            shares = [(1.0 / num_active_stats) for _ in range(num_active_stats)]
            if dist_mode == 2 and num_active_stats > 1:
                shares = [max(0.01, s + random.uniform(-skew_factor / 100.0, skew_factor / 100.0)) for s in shares]
                s_sum = sum(shares)
                shares = [s / s_sum for s in shares]
            
            allocated_values = [max(1, int(final_budget * s)) for s in shares]
            remainder = final_budget - sum(allocated_values)
            if remainder != 0 and allocated_values:
                max_idx = allocated_values.index(max(allocated_values))
                allocated_values[max_idx] += remainder
            
            for idx, stat_type in enumerate(chosen_stats):
                stats[f"stat_type{idx+1}"] = stat_type
                stats[f"stat_value{idx+1}"] = allocated_values[idx]
        
        m_key = (category["class"], category["subclass"], quality_code) 
        if m_key in material_library:
    # Pick a random valid pair from existing items in the database
         chosen_pair = random.choice(material_library[m_key])
         item_material = chosen_pair['Material']
         item_sheath = get_sheathe_type(category)
        else:
         fallback_key = (category["class"], category["subclass"])
         possible_keys = [k for k in material_library.keys() if k[0:2] == fallback_key]
         if possible_keys:
          chosen_pair = random.choice(random.choice([material_library[k] for k in possible_keys]))
          item_material = chosen_pair['Material']
          item_sheath = get_sheathe_type(category)
         else:
             item_material = 1
             item_sheath = 0
        cat_keys = (category["class"], category["subclass"])
        generated_name = generate_item_name((cat_keys[0], cat_keys[1], lookup_key[2]))
        display_obj = get_appropriate_display_id(cat_keys, ilvl, quality_code, category["InventoryType"])
        predicted_display_id = display_obj["id"]
        req_level = get_appropriate_req_level(cursor, ilvl, quality_code)
        if category['InventoryType'] in (2, 11, 23): avg_armor = 0.0  # Neck, Ring, Off-hand frill
        else: avg_armor = get_scaled_armor(lookup_database, category, quality_code, ilvl, sheet) if category['class'] == 4 else 0.0
        fuzz_factor = random.uniform(1.0 - variance, 1.0 + variance)
        final_armor = int(avg_armor * fuzz_factor) if avg_armor > 0 else 0
        final_block = max(1, int(get_scaled_block(lookup_database, category, quality_code, ilvl, sheet) * fuzz_factor)) if category['InventoryType'] == 14 else 0
        if bonding_mode == 1:   item_bonding = 2
        elif bonding_mode == 2: item_bonding = 1
        else:                   item_bonding = 2 if random.random() < boe_ratio else 1
        item_rand_prop = 0
        item_rand_suf  = 0
        if rand_mode == 2 and _ALL_ENCHANT_ENTRIES and random.random() < rand_pct:
            ench_entry = random.choice(_ALL_ENCHANT_ENTRIES)
            if ilvl < 20:
                item_rand_prop = ench_entry
            else:
                item_rand_suf  = ench_entry
        base_sell_price = calculate_item_sell_price(lookup_database, category, quality_code, ilvl)
        final_sell_price = int(base_sell_price * fuzz_factor)

        item_spellid = 0
        item_spelltrigger = 0
        item_spellcharges = 0
        item_spellcooldown = -1
        if spell_mode in (2, 3) and random.random() < (spell_pct / 100.0):
            item_spellid = batch_spell_id
            item_spelltrigger = batch_spell_trigger
            item_spellcharges = batch_spell_charges
            item_spellcooldown = batch_spell_cooldown

        internal_memory.append({
            "config": category, "quality": quality_code, "ilvl": ilvl,
            "name": generated_name, "displayid": display_obj.get("id"),
            "display_source": display_obj, "delay": dynamic_delay,
            "dmg_min": dmg_min, "dmg_max": dmg_max, "dps": final_dps,
            "armor": final_armor, "block": final_block,
            "stats": stats, "budget": final_budget,
            "required_level": req_level, "Material": item_material,
            "sheath": item_sheath, "sell_price": final_sell_price,
            "bonding": item_bonding,
            "rand_prop": item_rand_prop, "rand_suf": item_rand_suf,
            "itemset": 0,
            "spellid": item_spellid, "spelltrigger": item_spelltrigger,
            "spellcharges": item_spellcharges, "spellcooldown": item_spellcooldown
        })

    print(f"Saved {quantity} items. Total cached: {len(internal_memory)}")
    if get_input("\nAdd another batch? (y/n): ", lambda x: x.lower() if x.lower() in ['y', 'n'] else int("err")) == 'n': break

output_filename = "interactive_generated_items.sql"
conn = get_db_connection()
cursor = conn.cursor()
start_entry_id = get_next_entry_id(cursor)

print(f"\n💾 Session completed! Exporting to SQL database format...")
with open(output_filename, "w") as f:
    f.write("-- AI-Assisted Smart Item Generation Suite (Cohesive Synergy & Complete Spec Blueprints)\n\n")
    for idx, item in enumerate(internal_memory):
        current_id = start_entry_id + idx
        c = item["config"]
        s = item["stats"]
        description = f"iLvl {item['ilvl']} {c['name']} generated via structural blueprint routing."
        
        # Apostrophes in genitive names (e.g. Jin'do's) must be doubled for SQL.
        # The name in internal_memory is unchanged; only the SQL output escapes it.
        sql_name = item['name'].replace("'", "''")
        sql_string = f"""DELETE FROM `item_template` WHERE `entry` = {current_id};
INSERT INTO `item_template` (`entry`, `class`, `subclass`, `name`, `displayid`, `Quality`, `InventoryType`, `itemlevel`, `RequiredLevel`, `armor`, `block`, `bonding`, `RandomProperty`, `RandomSuffix`, `delay`, `dmg_min1`, `dmg_max1`, `dmg_type1`, `stat_type1`, `stat_value1`, `stat_type2`, `stat_value2`, `stat_type3`, `stat_value3`, `stat_type4`, `stat_value4`, `stat_type5`, `stat_value5`, `stat_type6`, `stat_value6`, `Material`, `sheath`, `SellPrice`, `itemset`, `spellid_1`, `spelltrigger_1`, `spellcharges_1`, `spellcooldown_1`, `spellcategory_1`, `spellcategorycooldown_1`, `Description`) 
VALUES ({current_id}, {c['class']}, {c['subclass']}, '{sql_name}', {item['displayid']}, {item['quality']}, {c['InventoryType']}, {item['ilvl']}, {item['required_level']}, {item.get('armor', 0)}, {item.get('block', 0)}, {item.get('bonding', 1)}, {item.get('rand_prop', 0)}, {item.get('rand_suf', 0)}, {item['delay']}, {item['dmg_min']}, {item['dmg_max']}, {c['dmg_type1']}, {s['stat_type1']}, {s['stat_value1']}, {s['stat_type2']}, {s['stat_value2']}, {s['stat_type3']}, {s['stat_value3']}, {s['stat_type4']}, {s['stat_value4']}, {s['stat_type5']}, {s['stat_value5']}, {s['stat_type6']}, {s['stat_value6']}, {item['Material']}, {item['sheath']}, {item.get('sell_price', 0)}, {item.get('itemset', 0)}, {item.get('spellid', 0)}, {item.get('spelltrigger', 0)}, {item.get('spellcharges', 0)}, {item.get('spellcooldown', -1)}, 0, -1, '{description}');\n"""
        f.write(sql_string)
        
        combat_info = f" ({item['dps']} DPS | Min-Max: {item['dmg_min']}-{item['dmg_max']})" if c['class'] == 2 else " (Armor Piece)"
        print(f" -> Compiled [ID: {current_id}] '{item['name']}' - iLvl {item['ilvl']}{combat_info}")
        
        primaries, secondaries = [], []
        for i in range(1, 7):
            st, sv = s[f"stat_type{i}"], s[f"stat_value{i}"]
            if sv > 0:
                if st in STAT_NAMES:
                    name, classification = STAT_NAMES[st]
                    formatted_stat = f"+{sv} {name}"
                    primaries.append(formatted_stat) if "Primary" in classification else secondaries.append(formatted_stat)
                else:
                    secondaries.append(f"+{sv} Unknown Stat (ID: {st})")
        
      
        print(f"    ↳ [Debug Allocation] Total Budget: {item['budget']}")
            
        # UPDATED: Print Source Metadata including the ID
        src = item.get("display_source", {})
        print(f"    ↳ [Visuals] DisplayID: {item['displayid']} (Source ID: {src.get('source_id', 'N/A')} | iLvl {src.get('lvl', 0)}, Qlty {src.get('q', 0)})")
            
        print(f"        🔹 Primaries:    [{', '.join(primaries) if primaries else 'None'}]")
        print(f"        🔸 Secondaries: [{', '.join(secondaries) if secondaries else 'None'}]")
         
cursor.close()
conn.close()
print(f"\n✅ Database connection closed. SQL file saved as {output_filename}")

csv_filename = "generated_items.csv"
print(f"\n💾 Exporting to CSV format: {csv_filename}")

import csv
with open(csv_filename, "w", newline='') as f:
    writer = csv.writer(f, quoting=csv.QUOTE_ALL)
    
    # Write the header row
    writer.writerow(["ID", "ClassID", "SubclassID", "Sound_Override_Subclassid", "Material", "DisplayInfoID", "InventoryType", "SheatheType"])
    
    # Write the data rows
    for idx, item in enumerate(internal_memory):
        current_id = start_entry_id + idx
        c = item["config"]
        
        writer.writerow([
            current_id,                      # ID
            c['class'],                      # ClassID
            c['subclass'],                   # SubclassID
            -1,                              # Sound_Override_Subclassid
            item.get('Material', 1),         # Material
            item['displayid'],               # DisplayInfoID
            c['InventoryType'],              # InventoryType
            item.get('sheath', 0)            # SheatheType
        ])

print(f"✅ CSV export successful!")

if generated_itemsets:
    itemset_csv_filename = "generated_itemsets.csv"
    print(f"\n💾 Exporting {len(generated_itemsets)} item set(s) to {itemset_csv_filename} (ItemSet.dbc format)...")

    LOCALE_SUFFIXES = ["enUS", "enGB", "koKR", "frFR", "deDE", "enCN", "zhCN",
                       "enTW", "zhTW", "esES", "esMX", "ruRU", "ptPT", "ptBR",
                       "itIT", "Unk"]  # 16 locale slots, WDBXEditor/AC convention

    with open(itemset_csv_filename, "w", newline='') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_ALL)
        header = ["ID"]
        header += [f"Name_Lang_{loc}" for loc in LOCALE_SUFFIXES]
        header += ["Name_Lang_Mask"]
        header += [f"ItemID_{i}" for i in range(1, 18)]
        header += [f"SetSpellID_{i}" for i in range(1, 9)]
        header += [f"SetThreshold_{i}" for i in range(1, 9)]
        header += ["RequiredSkill", "RequiredSkillRank"]
        writer.writerow(header)

        for iset in generated_itemsets:
            entry_ids = [start_entry_id + idx for idx in iset["item_indices"]]
            item_id_cols = entry_ids + [0] * (17 - len(entry_ids))
            spell_cols = [s for (_, s) in iset["spells"]] + [0] * (8 - len(iset["spells"]))
            threshold_cols = [t for (t, _) in iset["spells"]] + [0] * (8 - len(iset["spells"]))
            # Only enUS gets the name; every other locale stays empty (client falls
            # back to enUS via the mask, same convention AC uses for its own DBCs).
            name_cols = [iset["name"]] + [""] * (len(LOCALE_SUFFIXES) - 1)
            row = ([iset["id"]] + name_cols + [16712190]
                    + item_id_cols + spell_cols + threshold_cols
                    + [0, 0])  # RequiredSkill, RequiredSkillRank
            writer.writerow(row)
            print(f"   -> Set [{iset['id']}] '{iset['name']}' -- items: {entry_ids}")
    print(f"✅ Item Set export successful!")

print("🏁 Generation complete! compiling Excel tooltips sheet...")

# ⚡ Call the tooltips function directly here without any 'except' statement:
export_to_excel_tooltips(internal_memory, filename="generated_items_tooltips.xlsx")
